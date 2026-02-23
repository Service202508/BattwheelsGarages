"""
Financial Reports Module
Provides comprehensive financial reports with PDF and Excel export capabilities
- Profit & Loss Statement
- Balance Sheet
- Accounts Receivable Aging
- Accounts Payable Aging
- Sales by Customer Report
"""
from fastapi import APIRouter, HTTPException, Query, Request, Depends
from fastapi.responses import Response, StreamingResponse
from datetime import datetime, timezone, timedelta
from motor.motor_asyncio import AsyncIOMotorClient
from io import BytesIO
import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from core.subscriptions.entitlement import require_feature

# Soft import for PDF service (may not be available in all environments)
PDF_SERVICE_AVAILABLE = False
generate_pdf_from_html = None
try:
    from services.pdf_service import generate_pdf_from_html
    PDF_SERVICE_AVAILABLE = True
except Exception:
    pass

router = APIRouter(prefix="/reports", tags=["Financial Reports"])

# Database connection
def get_db():
    mongo_url = os.environ['MONGO_URL']
    client = AsyncIOMotorClient(mongo_url)
    return client[os.environ['DB_NAME']]

# ============== HELPER FUNCTIONS ==============

def parse_date(date_str: str, default_days_ago: int = 365) -> datetime:
    """Parse date string or return default"""
    if date_str:
        try:
            return datetime.strptime(date_str, '%Y-%m-%d').replace(tzinfo=timezone.utc)
        except ValueError:
            pass
    return datetime.now(timezone.utc) - timedelta(days=default_days_ago)

def format_currency(value: float) -> str:
    """Format currency in Indian Rupee format"""
    return f"₹{value:,.2f}"

# ============== PDF GENERATORS ==============

def generate_profit_loss_html(data: dict, org_settings: dict = None) -> str:
    """Generate HTML for Profit & Loss report"""
    org = org_settings or {}
    company_name = org.get('company_name', 'Battwheels')
    
    expenses_html = ""
    for cat, amt in data.get('expenses_breakdown', {}).items():
        expenses_html += f"""
            <tr>
                <td style="padding-left: 30px;">{cat}</td>
                <td class="amount">{format_currency(amt)}</td>
            </tr>
        """
    
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            @page {{ size: A4; margin: 1cm; }}
            body {{
                font-family: 'Helvetica Neue', Arial, sans-serif;
                font-size: 10pt;
                color: #333;
                margin: 0;
                padding: 20px;
            }}
            .header {{
                text-align: center;
                border-bottom: 3px solid #22EDA9;
                padding-bottom: 15px;
                margin-bottom: 20px;
            }}
            .company-name {{ font-size: 20pt; font-weight: bold; color: #1a1a1a; }}
            .report-title {{ font-size: 16pt; color: #22EDA9; margin-top: 10px; }}
            .period {{ font-size: 10pt; color: #666; margin-top: 5px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
            th {{ background: #f8f9fa; padding: 12px 8px; text-align: left; font-weight: bold; border-bottom: 2px solid #22EDA9; }}
            td {{ padding: 10px 8px; border-bottom: 1px solid #eee; }}
            .section-header {{ background: #f0f0f0; font-weight: bold; }}
            .section-total {{ font-weight: bold; background: #fafafa; }}
            .grand-total {{ font-weight: bold; font-size: 12pt; background: #22EDA9; color: #000; }}
            .amount {{ text-align: right; }}
            .positive {{ color: #28a745; }}
            .negative {{ color: #dc3545; }}
            .footer {{ margin-top: 30px; text-align: center; font-size: 8pt; color: #999; }}
        </style>
    </head>
    <body>
        <div class="header">
            <div class="company-name">{company_name}</div>
            <div class="report-title">PROFIT & LOSS STATEMENT</div>
            <div class="period">{data['period']['start_date']} to {data['period']['end_date']}</div>
        </div>
        
        <table>
            <thead>
                <tr>
                    <th>Account</th>
                    <th class="amount">Amount</th>
                </tr>
            </thead>
            <tbody>
                <tr class="section-header">
                    <td colspan="2">INCOME</td>
                </tr>
                <tr>
                    <td style="padding-left: 20px;">Operating Income (Sales)</td>
                    <td class="amount">{format_currency(data['total_income'])}</td>
                </tr>
                <tr class="section-total">
                    <td>Total Income</td>
                    <td class="amount">{format_currency(data['total_income'])}</td>
                </tr>
                
                <tr class="section-header">
                    <td colspan="2">COST OF GOODS SOLD</td>
                </tr>
                <tr>
                    <td style="padding-left: 20px;">Direct Costs (Purchases/Bills)</td>
                    <td class="amount">{format_currency(data['total_cogs'])}</td>
                </tr>
                <tr class="section-total">
                    <td>Gross Profit</td>
                    <td class="amount {'positive' if data['gross_profit'] >= 0 else 'negative'}">{format_currency(data['gross_profit'])}</td>
                </tr>
                
                <tr class="section-header">
                    <td colspan="2">OPERATING EXPENSES</td>
                </tr>
                {expenses_html}
                <tr class="section-total">
                    <td>Total Expenses</td>
                    <td class="amount">{format_currency(data['total_expenses'])}</td>
                </tr>
                
                <tr class="grand-total">
                    <td>NET PROFIT</td>
                    <td class="amount {'positive' if data['net_profit'] >= 0 else 'negative'}">{format_currency(data['net_profit'])}</td>
                </tr>
            </tbody>
        </table>
        
        <div class="footer">
            Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')} | {company_name}
        </div>
    </body>
    </html>
    """
    return html

def generate_balance_sheet_html(data: dict, org_settings: dict = None) -> str:
    """Generate HTML for Balance Sheet report"""
    org = org_settings or {}
    company_name = org.get('company_name', 'Battwheels')
    
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            @page {{ size: A4; margin: 1cm; }}
            body {{
                font-family: 'Helvetica Neue', Arial, sans-serif;
                font-size: 10pt;
                color: #333;
                margin: 0;
                padding: 20px;
            }}
            .header {{
                text-align: center;
                border-bottom: 3px solid #22EDA9;
                padding-bottom: 15px;
                margin-bottom: 20px;
            }}
            .company-name {{ font-size: 20pt; font-weight: bold; color: #1a1a1a; }}
            .report-title {{ font-size: 16pt; color: #22EDA9; margin-top: 10px; }}
            .as-of {{ font-size: 10pt; color: #666; margin-top: 5px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
            th {{ background: #f8f9fa; padding: 12px 8px; text-align: left; font-weight: bold; border-bottom: 2px solid #22EDA9; }}
            td {{ padding: 10px 8px; border-bottom: 1px solid #eee; }}
            .section-header {{ background: #e8f5f0; font-weight: bold; color: #1a6b47; }}
            .subsection {{ padding-left: 20px !important; }}
            .section-total {{ font-weight: bold; background: #fafafa; }}
            .grand-total {{ font-weight: bold; font-size: 12pt; background: #22EDA9; color: #000; }}
            .amount {{ text-align: right; }}
            .footer {{ margin-top: 30px; text-align: center; font-size: 8pt; color: #999; }}
        </style>
    </head>
    <body>
        <div class="header">
            <div class="company-name">{company_name}</div>
            <div class="report-title">BALANCE SHEET</div>
            <div class="as-of">As of {data['as_of_date']}</div>
        </div>
        
        <table>
            <thead>
                <tr>
                    <th>Account</th>
                    <th class="amount">Amount</th>
                </tr>
            </thead>
            <tbody>
                <tr class="section-header">
                    <td colspan="2">ASSETS</td>
                </tr>
                <tr>
                    <td class="subsection">Accounts Receivable</td>
                    <td class="amount">{format_currency(data['assets']['accounts_receivable'])}</td>
                </tr>
                <tr>
                    <td class="subsection">Bank Balance</td>
                    <td class="amount">{format_currency(data['assets']['bank_balance'])}</td>
                </tr>
                <tr>
                    <td class="subsection">Inventory Value</td>
                    <td class="amount">{format_currency(data['assets']['inventory_value'])}</td>
                </tr>
                <tr class="section-total">
                    <td>Total Assets</td>
                    <td class="amount">{format_currency(data['assets']['total'])}</td>
                </tr>
                
                <tr class="section-header">
                    <td colspan="2">LIABILITIES</td>
                </tr>
                <tr>
                    <td class="subsection">Accounts Payable</td>
                    <td class="amount">{format_currency(data['liabilities']['accounts_payable'])}</td>
                </tr>
                <tr class="section-total">
                    <td>Total Liabilities</td>
                    <td class="amount">{format_currency(data['liabilities']['total'])}</td>
                </tr>
                
                <tr class="section-header">
                    <td colspan="2">EQUITY</td>
                </tr>
                <tr>
                    <td class="subsection">Retained Earnings</td>
                    <td class="amount">{format_currency(data['equity']['retained_earnings'])}</td>
                </tr>
                <tr class="grand-total">
                    <td>Total Equity</td>
                    <td class="amount">{format_currency(data['equity']['total'])}</td>
                </tr>
            </tbody>
        </table>
        
        <div class="footer">
            Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')} | {company_name}
        </div>
    </body>
    </html>
    """
    return html

def generate_ar_aging_html(data: dict, org_settings: dict = None) -> str:
    """Generate HTML for AR Aging report"""
    org = org_settings or {}
    company_name = org.get('company_name', 'Battwheels')
    
    aging_data = data.get('aging_data', {})
    invoices_html = ""
    for inv in data.get('invoices', [])[:50]:  # Limit to 50 for PDF
        invoices_html += f"""
            <tr>
                <td>{inv.get('invoice_number', '-')}</td>
                <td>{inv.get('customer_name', '-')}</td>
                <td>{inv.get('due_date', '-')}</td>
                <td>{inv.get('days_overdue', 0)}</td>
                <td class="amount">{format_currency(inv.get('balance', 0))}</td>
            </tr>
        """
    
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            @page {{ size: A4 landscape; margin: 1cm; }}
            body {{
                font-family: 'Helvetica Neue', Arial, sans-serif;
                font-size: 9pt;
                color: #333;
                margin: 0;
                padding: 20px;
            }}
            .header {{
                text-align: center;
                border-bottom: 3px solid #22EDA9;
                padding-bottom: 15px;
                margin-bottom: 20px;
            }}
            .company-name {{ font-size: 18pt; font-weight: bold; }}
            .report-title {{ font-size: 14pt; color: #22EDA9; margin-top: 10px; }}
            .aging-summary {{
                display: flex;
                justify-content: space-between;
                margin-bottom: 20px;
                gap: 10px;
            }}
            .aging-box {{
                flex: 1;
                text-align: center;
                padding: 15px;
                background: #f8f9fa;
                border-radius: 8px;
                border: 1px solid #eee;
            }}
            .aging-box.current {{ border-left: 4px solid #28a745; }}
            .aging-box.warning {{ border-left: 4px solid #ffc107; }}
            .aging-box.danger {{ border-left: 4px solid #dc3545; }}
            .aging-label {{ font-size: 9pt; color: #666; }}
            .aging-amount {{ font-size: 14pt; font-weight: bold; margin-top: 5px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
            th {{ background: #f8f9fa; padding: 10px 6px; text-align: left; font-size: 9pt; border-bottom: 2px solid #22EDA9; }}
            td {{ padding: 8px 6px; border-bottom: 1px solid #eee; }}
            .amount {{ text-align: right; }}
            .grand-total {{ font-weight: bold; background: #22EDA9; }}
            .footer {{ margin-top: 20px; text-align: center; font-size: 8pt; color: #999; }}
        </style>
    </head>
    <body>
        <div class="header">
            <div class="company-name">{company_name}</div>
            <div class="report-title">ACCOUNTS RECEIVABLE AGING</div>
            <div style="font-size: 10pt; color: #666;">As of {data['as_of_date']}</div>
        </div>
        
        <div class="aging-summary">
            <div class="aging-box current">
                <div class="aging-label">Current</div>
                <div class="aging-amount">{format_currency(aging_data.get('current', 0))}</div>
            </div>
            <div class="aging-box">
                <div class="aging-label">1-30 Days</div>
                <div class="aging-amount">{format_currency(aging_data.get('1_30', 0))}</div>
            </div>
            <div class="aging-box warning">
                <div class="aging-label">31-60 Days</div>
                <div class="aging-amount">{format_currency(aging_data.get('31_60', 0))}</div>
            </div>
            <div class="aging-box warning">
                <div class="aging-label">61-90 Days</div>
                <div class="aging-amount">{format_currency(aging_data.get('61_90', 0))}</div>
            </div>
            <div class="aging-box danger">
                <div class="aging-label">Over 90 Days</div>
                <div class="aging-amount">{format_currency(aging_data.get('over_90', 0))}</div>
            </div>
        </div>
        
        <table>
            <thead>
                <tr>
                    <th>Invoice #</th>
                    <th>Customer</th>
                    <th>Due Date</th>
                    <th>Days Overdue</th>
                    <th class="amount">Balance</th>
                </tr>
            </thead>
            <tbody>
                {invoices_html}
                <tr class="grand-total">
                    <td colspan="4">Total Outstanding</td>
                    <td class="amount">{format_currency(data['total_ar'])}</td>
                </tr>
            </tbody>
        </table>
        
        <div class="footer">
            Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')} | {company_name}
        </div>
    </body>
    </html>
    """
    return html

def generate_sales_by_customer_html(data: dict, org_settings: dict = None) -> str:
    """Generate HTML for Sales by Customer report"""
    org = org_settings or {}
    company_name = org.get('company_name', 'Battwheels')
    
    rows_html = ""
    for idx, item in enumerate(data.get('sales_data', []), 1):
        rows_html += f"""
            <tr>
                <td>{idx}</td>
                <td>{item.get('customer_name', 'Unknown')}</td>
                <td class="amount">{item.get('invoice_count', 0)}</td>
                <td class="amount">{format_currency(item.get('total_sales', 0))}</td>
            </tr>
        """
    
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            @page {{ size: A4; margin: 1cm; }}
            body {{
                font-family: 'Helvetica Neue', Arial, sans-serif;
                font-size: 10pt;
                color: #333;
                margin: 0;
                padding: 20px;
            }}
            .header {{
                text-align: center;
                border-bottom: 3px solid #22EDA9;
                padding-bottom: 15px;
                margin-bottom: 20px;
            }}
            .company-name {{ font-size: 20pt; font-weight: bold; }}
            .report-title {{ font-size: 16pt; color: #22EDA9; margin-top: 10px; }}
            .period {{ font-size: 10pt; color: #666; margin-top: 5px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
            th {{ background: #f8f9fa; padding: 12px 8px; text-align: left; font-weight: bold; border-bottom: 2px solid #22EDA9; }}
            td {{ padding: 10px 8px; border-bottom: 1px solid #eee; }}
            .amount {{ text-align: right; }}
            .grand-total {{ font-weight: bold; font-size: 12pt; background: #22EDA9; color: #000; }}
            .footer {{ margin-top: 30px; text-align: center; font-size: 8pt; color: #999; }}
        </style>
    </head>
    <body>
        <div class="header">
            <div class="company-name">{company_name}</div>
            <div class="report-title">SALES BY CUSTOMER</div>
            <div class="period">{data['period']['start_date']} to {data['period']['end_date']}</div>
        </div>
        
        <table>
            <thead>
                <tr>
                    <th style="width: 50px;">#</th>
                    <th>Customer Name</th>
                    <th class="amount">Invoices</th>
                    <th class="amount">Total Sales</th>
                </tr>
            </thead>
            <tbody>
                {rows_html}
                <tr class="grand-total">
                    <td colspan="2">TOTAL</td>
                    <td class="amount">{data.get('total_invoices', 0)}</td>
                    <td class="amount">{format_currency(data.get('total_sales', 0))}</td>
                </tr>
            </tbody>
        </table>
        
        <div class="footer">
            Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')} | {company_name}
        </div>
    </body>
    </html>
    """
    return html

# ============== EXCEL GENERATORS ==============

def create_styled_workbook():
    """Create a workbook with standard styles"""
    wb = openpyxl.Workbook()
    return wb

def style_header_row(ws, row_num, num_cols):
    """Apply header styling to a row"""
    header_fill = PatternFill(start_color="22EDA9", end_color="22EDA9", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    header_border = Border(
        bottom=Side(style='medium', color='000000')
    )
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = Alignment(horizontal='center')

def style_total_row(ws, row_num, num_cols):
    """Apply total row styling"""
    total_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    total_font = Font(bold=True)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = total_fill
        cell.font = total_font

def generate_profit_loss_excel(data: dict) -> bytes:
    """Generate Excel for Profit & Loss report"""
    wb = create_styled_workbook()
    ws = wb.active
    ws.title = "Profit & Loss"
    
    # Title
    ws.append([f"Profit & Loss Statement"])
    ws.merge_cells('A1:B1')
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([f"Period: {data['period']['start_date']} to {data['period']['end_date']}"])
    ws.append([])
    
    # Headers
    ws.append(['Account', 'Amount'])
    style_header_row(ws, 4, 2)
    
    # Income
    ws.append(['INCOME', ''])
    ws['A5'].font = Font(bold=True)
    ws.append(['  Operating Income', data['total_income']])
    ws.append(['Total Income', data['total_income']])
    style_total_row(ws, 7, 2)
    
    # COGS
    ws.append([])
    ws.append(['COST OF GOODS SOLD', ''])
    ws['A9'].font = Font(bold=True)
    ws.append(['  Direct Costs', data['total_cogs']])
    ws.append(['Gross Profit', data['gross_profit']])
    style_total_row(ws, 11, 2)
    
    # Operating Expenses
    ws.append([])
    ws.append(['OPERATING EXPENSES', ''])
    ws['A13'].font = Font(bold=True)
    row = 14
    for cat, amt in data.get('expenses_breakdown', {}).items():
        ws.append([f"  {cat}", amt])
        row += 1
    ws.append(['Total Expenses', data['total_expenses']])
    style_total_row(ws, row, 2)
    
    # Net Profit
    ws.append([])
    ws.append(['NET PROFIT', data['net_profit']])
    last_row = ws.max_row
    ws[f'A{last_row}'].font = Font(bold=True, size=12)
    ws[f'B{last_row}'].font = Font(bold=True, size=12)
    
    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    
    # Format numbers
    for row in ws.iter_rows(min_row=5, max_col=2):
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.value != '':
                cell.number_format = '₹#,##0.00'
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def generate_balance_sheet_excel(data: dict) -> bytes:
    """Generate Excel for Balance Sheet report"""
    wb = create_styled_workbook()
    ws = wb.active
    ws.title = "Balance Sheet"
    
    ws.append(['Balance Sheet'])
    ws.merge_cells('A1:B1')
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([f"As of: {data['as_of_date']}"])
    ws.append([])
    
    ws.append(['Account', 'Amount'])
    style_header_row(ws, 4, 2)
    
    # Assets
    ws.append(['ASSETS', ''])
    ws['A5'].font = Font(bold=True)
    ws.append(['  Accounts Receivable', data['assets']['accounts_receivable']])
    ws.append(['  Bank Balance', data['assets']['bank_balance']])
    ws.append(['  Inventory Value', data['assets']['inventory_value']])
    ws.append(['Total Assets', data['assets']['total']])
    style_total_row(ws, 9, 2)
    
    # Liabilities
    ws.append([])
    ws.append(['LIABILITIES', ''])
    ws['A11'].font = Font(bold=True)
    ws.append(['  Accounts Payable', data['liabilities']['accounts_payable']])
    ws.append(['Total Liabilities', data['liabilities']['total']])
    style_total_row(ws, 13, 2)
    
    # Equity
    ws.append([])
    ws.append(['EQUITY', ''])
    ws['A15'].font = Font(bold=True)
    ws.append(['  Retained Earnings', data['equity']['retained_earnings']])
    ws.append(['Total Equity', data['equity']['total']])
    style_total_row(ws, 17, 2)
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    
    for row in ws.iter_rows(min_row=5, max_col=2):
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.value != '':
                cell.number_format = '₹#,##0.00'
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def generate_ar_aging_excel(data: dict) -> bytes:
    """Generate Excel for AR Aging report"""
    wb = create_styled_workbook()
    ws = wb.active
    ws.title = "AR Aging"
    
    ws.append(['Accounts Receivable Aging Report'])
    ws.merge_cells('A1:E1')
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([f"As of: {data['as_of_date']}"])
    ws.append([])
    
    # Aging Summary
    aging_data = data.get('aging_data', {})
    ws.append(['Aging Summary'])
    ws['A4'].font = Font(bold=True)
    ws.append(['Current', '1-30 Days', '31-60 Days', '61-90 Days', 'Over 90 Days'])
    style_header_row(ws, 5, 5)
    ws.append([
        aging_data.get('current', 0),
        aging_data.get('1_30', 0),
        aging_data.get('31_60', 0),
        aging_data.get('61_90', 0),
        aging_data.get('over_90', 0)
    ])
    for col in range(1, 6):
        ws.cell(row=6, column=col).number_format = '₹#,##0.00'
    
    ws.append([])
    ws.append(['Invoice Details'])
    ws['A8'].font = Font(bold=True)
    
    ws.append(['Invoice Number', 'Customer', 'Due Date', 'Days Overdue', 'Balance'])
    style_header_row(ws, 9, 5)
    
    for inv in data.get('invoices', []):
        ws.append([
            inv.get('invoice_number', '-'),
            inv.get('customer_name', '-'),
            inv.get('due_date', '-'),
            inv.get('days_overdue', 0),
            inv.get('balance', 0)
        ])
    
    ws.append(['Total', '', '', '', data['total_ar']])
    style_total_row(ws, ws.max_row, 5)
    
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    
    for row in ws.iter_rows(min_row=10, max_col=5):
        if row[4].value and isinstance(row[4].value, (int, float)):
            row[4].number_format = '₹#,##0.00'
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def generate_sales_by_customer_excel(data: dict) -> bytes:
    """Generate Excel for Sales by Customer report"""
    wb = create_styled_workbook()
    ws = wb.active
    ws.title = "Sales by Customer"
    
    ws.append(['Sales by Customer Report'])
    ws.merge_cells('A1:D1')
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([f"Period: {data['period']['start_date']} to {data['period']['end_date']}"])
    ws.append([])
    
    ws.append(['#', 'Customer Name', 'Invoice Count', 'Total Sales'])
    style_header_row(ws, 4, 4)
    
    for idx, item in enumerate(data.get('sales_data', []), 1):
        ws.append([
            idx,
            item.get('customer_name', 'Unknown'),
            item.get('invoice_count', 0),
            item.get('total_sales', 0)
        ])
    
    ws.append(['', 'TOTAL', data.get('total_invoices', 0), data.get('total_sales', 0)])
    style_total_row(ws, ws.max_row, 4)
    
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    
    for row in ws.iter_rows(min_row=5, max_col=4):
        if row[3].value and isinstance(row[3].value, (int, float)):
            row[3].number_format = '₹#,##0.00'
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

# ============== API ENDPOINTS ==============

@router.get("/profit-loss")
async def get_profit_loss_report(
    start_date: str = "",
    end_date: str = "",
    format: str = Query("json", enum=["json", "pdf", "excel"]),
    request: Request = None,
    _: None = Depends(require_feature("advanced_reports"))
):
    """
    Profit & Loss Report (Income Statement)
    Aggregate income and expense accounts for a period.
    Exports: JSON (web view), PDF, Excel
    """
    db = get_db()
    
    if not start_date:
        start_date = (datetime.now(timezone.utc) - timedelta(days=365)).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    # Income: Sum from invoices
    income_pipeline = [
        {"$match": {"date": {"$gte": start_date, "$lte": end_date}}},
        {"$group": {"_id": None, "total": {"$sum": "$total"}}}
    ]
    income_result = await db.invoices.aggregate(income_pipeline).to_list(1)
    total_income = income_result[0]["total"] if income_result else 0
    
    # COGS: Sum from bills
    cogs_pipeline = [
        {"$match": {"date": {"$gte": start_date, "$lte": end_date}}},
        {"$group": {"_id": None, "total": {"$sum": "$total"}}}
    ]
    cogs_result = await db.bills.aggregate(cogs_pipeline).to_list(1)
    total_cogs = cogs_result[0]["total"] if cogs_result else 0
    
    gross_profit = total_income - total_cogs
    
    # Operating Expenses: Group by category
    expense_pipeline = [
        {"$match": {"expense_date": {"$gte": start_date, "$lte": end_date}}},
        {"$group": {"_id": "$expense_account", "total": {"$sum": "$amount"}}}
    ]
    expenses_result = await db.expenses.aggregate(expense_pipeline).to_list(100)
    expenses_breakdown = {e["_id"] or "Uncategorized": e["total"] for e in expenses_result}
    total_expenses = sum(expenses_breakdown.values())
    
    net_profit = gross_profit - total_expenses
    
    report_data = {
        "period": {"start_date": start_date, "end_date": end_date},
        "total_income": total_income,
        "total_cogs": total_cogs,
        "gross_profit": gross_profit,
        "expenses_breakdown": expenses_breakdown,
        "total_expenses": total_expenses,
        "net_profit": net_profit,
        "margins": {
            "gross_margin_percent": round((gross_profit / total_income * 100), 2) if total_income > 0 else 0,
            "net_margin_percent": round((net_profit / total_income * 100), 2) if total_income > 0 else 0
        }
    }
    
    if format == "pdf":
        org_settings = await db.organization_settings.find_one({}, {"_id": 0}) or {}
        html_content = generate_profit_loss_html(report_data, org_settings)
        pdf_bytes = generate_pdf_from_html(html_content)
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=profit_loss_{start_date}_to_{end_date}.pdf"}
        )
    
    elif format == "excel":
        excel_bytes = generate_profit_loss_excel(report_data)
        return Response(
            content=excel_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=profit_loss_{start_date}_to_{end_date}.xlsx"}
        )
    
    return {"code": 0, "report": "profit_and_loss", **report_data}

@router.get("/balance-sheet")
async def get_balance_sheet_report(
    as_of_date: str = "",
    format: str = Query("json", enum=["json", "pdf", "excel"])
):
    """
    Balance Sheet Report
    Aggregate assets, liabilities, equity as of a date.
    Exports: JSON (web view), PDF, Excel
    """
    db = get_db()
    
    if not as_of_date:
        as_of_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    # Assets - Accounts Receivable
    receivables_pipeline = [
        {"$match": {"balance": {"$gt": 0}}},
        {"$group": {"_id": None, "total": {"$sum": "$balance"}}}
    ]
    receivables = await db.invoices.aggregate(receivables_pipeline).to_list(1)
    accounts_receivable = receivables[0]["total"] if receivables else 0
    
    # Assets - Bank Balance
    bank_pipeline = [
        {"$match": {"is_active": True}},
        {"$group": {"_id": None, "total": {"$sum": "$balance"}}}
    ]
    bank_result = await db.bankaccounts.aggregate(bank_pipeline).to_list(1)
    bank_balance = bank_result[0]["total"] if bank_result else 0
    
    # Assets - Inventory Value
    inventory_pipeline = [
        {"$match": {"status": "active"}},
        {"$group": {"_id": None, "total": {"$sum": {"$multiply": ["$stock_on_hand", "$rate"]}}}}
    ]
    inventory_result = await db.items.aggregate(inventory_pipeline).to_list(1)
    inventory_value = inventory_result[0]["total"] if inventory_result else 0
    
    # Liabilities - Accounts Payable
    payables_pipeline = [
        {"$match": {"balance": {"$gt": 0}}},
        {"$group": {"_id": None, "total": {"$sum": "$balance"}}}
    ]
    payables = await db.bills.aggregate(payables_pipeline).to_list(1)
    accounts_payable = payables[0]["total"] if payables else 0
    
    total_assets = accounts_receivable + bank_balance + inventory_value
    total_liabilities = accounts_payable
    equity = total_assets - total_liabilities
    
    report_data = {
        "as_of_date": as_of_date,
        "assets": {
            "accounts_receivable": accounts_receivable,
            "bank_balance": bank_balance,
            "inventory_value": inventory_value,
            "total": total_assets
        },
        "liabilities": {
            "accounts_payable": accounts_payable,
            "total": total_liabilities
        },
        "equity": {
            "retained_earnings": equity,
            "total": equity
        }
    }
    
    if format == "pdf":
        org_settings = await db.organization_settings.find_one({}, {"_id": 0}) or {}
        html_content = generate_balance_sheet_html(report_data, org_settings)
        pdf_bytes = generate_pdf_from_html(html_content)
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=balance_sheet_{as_of_date}.pdf"}
        )
    
    elif format == "excel":
        excel_bytes = generate_balance_sheet_excel(report_data)
        return Response(
            content=excel_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=balance_sheet_{as_of_date}.xlsx"}
        )
    
    return {"code": 0, "report": "balance_sheet", **report_data}

@router.get("/ar-aging")
async def get_ar_aging_report(
    as_of_date: str = "",
    format: str = Query("json", enum=["json", "pdf", "excel"])
):
    """
    Accounts Receivable Aging Report
    Group unpaid invoices by age buckets.
    Exports: JSON (web view), PDF, Excel
    """
    db = get_db()
    
    if not as_of_date:
        as_of_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    as_of_datetime = datetime.strptime(as_of_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    
    # Get unpaid invoices (limited for performance)
    invoices = await db.invoices.find(
        {"balance": {"$gt": 0}, "status": {"$in": ["sent", "partial", "overdue", "draft"]}},
        {"_id": 0}
    ).to_list(length=1000)
    
    aging_data = {
        "current": 0,
        "1_30": 0,
        "31_60": 0,
        "61_90": 0,
        "over_90": 0
    }
    
    invoice_details = []
    
    for inv in invoices:
        try:
            due_date = datetime.strptime(inv.get("due_date", ""), "%Y-%m-%d").replace(tzinfo=timezone.utc)
        except (ValueError, TypeError):
            continue
        
        days_overdue = (as_of_datetime - due_date).days
        balance = inv.get("balance", 0)
        
        inv_detail = {
            "invoice_number": inv.get("invoice_number", "-"),
            "customer_name": inv.get("customer_name", "-"),
            "due_date": inv.get("due_date", "-"),
            "days_overdue": max(0, days_overdue),
            "balance": balance
        }
        invoice_details.append(inv_detail)
        
        if days_overdue <= 0:
            aging_data["current"] += balance
        elif days_overdue <= 30:
            aging_data["1_30"] += balance
        elif days_overdue <= 60:
            aging_data["31_60"] += balance
        elif days_overdue <= 90:
            aging_data["61_90"] += balance
        else:
            aging_data["over_90"] += balance
    
    # Sort by days overdue descending
    invoice_details.sort(key=lambda x: x["days_overdue"], reverse=True)
    
    total_ar = sum(aging_data.values())
    
    report_data = {
        "as_of_date": as_of_date,
        "aging_data": aging_data,
        "total_ar": total_ar,
        "invoices": invoice_details
    }
    
    if format == "pdf":
        org_settings = await db.organization_settings.find_one({}, {"_id": 0}) or {}
        html_content = generate_ar_aging_html(report_data, org_settings)
        pdf_bytes = generate_pdf_from_html(html_content)
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=ar_aging_{as_of_date}.pdf"}
        )
    
    elif format == "excel":
        excel_bytes = generate_ar_aging_excel(report_data)
        return Response(
            content=excel_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=ar_aging_{as_of_date}.xlsx"}
        )
    
    return {"code": 0, "report": "ar_aging", **report_data}

@router.get("/sales-by-customer")
async def get_sales_by_customer_report(
    start_date: str = "",
    end_date: str = "",
    format: str = Query("json", enum=["json", "pdf", "excel"])
):
    """
    Sales by Customer Report
    Aggregate total sales per customer for a period.
    Exports: JSON (web view), PDF, Excel
    """
    db = get_db()
    
    if not start_date:
        start_date = (datetime.now(timezone.utc) - timedelta(days=365)).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    # Aggregate sales by customer
    sales_pipeline = [
        {"$match": {"date": {"$gte": start_date, "$lte": end_date}}},
        {"$group": {
            "_id": "$customer_name",
            "total_sales": {"$sum": "$total"},
            "invoice_count": {"$sum": 1}
        }},
        {"$sort": {"total_sales": -1}}
    ]
    
    sales_result = await db.invoices.aggregate(sales_pipeline).to_list(1000)
    
    sales_data = []
    total_sales = 0
    total_invoices = 0
    
    for item in sales_result:
        sales_data.append({
            "customer_name": item["_id"] or "Unknown",
            "total_sales": item["total_sales"],
            "invoice_count": item["invoice_count"]
        })
        total_sales += item["total_sales"]
        total_invoices += item["invoice_count"]
    
    report_data = {
        "period": {"start_date": start_date, "end_date": end_date},
        "sales_data": sales_data,
        "total_sales": total_sales,
        "total_invoices": total_invoices
    }
    
    if format == "pdf":
        org_settings = await db.organization_settings.find_one({}, {"_id": 0}) or {}
        html_content = generate_sales_by_customer_html(report_data, org_settings)
        pdf_bytes = generate_pdf_from_html(html_content)
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=sales_by_customer_{start_date}_to_{end_date}.pdf"}
        )
    
    elif format == "excel":
        excel_bytes = generate_sales_by_customer_excel(report_data)
        return Response(
            content=excel_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=sales_by_customer_{start_date}_to_{end_date}.xlsx"}
        )
    
    return {"code": 0, "report": "sales_by_customer", **report_data}

@router.get("/ap-aging")
async def get_ap_aging_report(
    as_of_date: str = "",
    format: str = Query("json", enum=["json", "pdf", "excel"])
):
    """
    Accounts Payable Aging Report
    Group unpaid bills by age buckets.
    Exports: JSON (web view), PDF, Excel
    """
    db = get_db()
    
    if not as_of_date:
        as_of_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    as_of_datetime = datetime.strptime(as_of_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    
    # Get unpaid bills
    bills = await db.bills.find(
        {"balance": {"$gt": 0}},
        {"_id": 0}
    ).to_list(length=10000)
    
    aging_data = {
        "current": 0,
        "1_30": 0,
        "31_60": 0,
        "61_90": 0,
        "over_90": 0
    }
    
    bill_details = []
    
    for bill in bills:
        try:
            due_date = datetime.strptime(bill.get("due_date", ""), "%Y-%m-%d").replace(tzinfo=timezone.utc)
        except (ValueError, TypeError):
            continue
        
        days_overdue = (as_of_datetime - due_date).days
        balance = bill.get("balance", 0)
        
        bill_detail = {
            "bill_number": bill.get("bill_number", "-"),
            "vendor_name": bill.get("vendor_name", "-"),
            "due_date": bill.get("due_date", "-"),
            "days_overdue": max(0, days_overdue),
            "balance": balance
        }
        bill_details.append(bill_detail)
        
        if days_overdue <= 0:
            aging_data["current"] += balance
        elif days_overdue <= 30:
            aging_data["1_30"] += balance
        elif days_overdue <= 60:
            aging_data["31_60"] += balance
        elif days_overdue <= 90:
            aging_data["61_90"] += balance
        else:
            aging_data["over_90"] += balance
    
    bill_details.sort(key=lambda x: x["days_overdue"], reverse=True)
    total_ap = sum(aging_data.values())
    
    report_data = {
        "as_of_date": as_of_date,
        "aging_data": aging_data,
        "total_ap": total_ap,
        "bills": bill_details
    }
    
    if format == "json":
        return {"code": 0, "report": "ap_aging", **report_data}
    
    # For PDF/Excel, reuse similar logic as AR aging
    elif format == "pdf":
        org_settings = await db.organization_settings.find_one({}, {"_id": 0}) or {}
        html_content = generate_ap_aging_html(report_data, org_settings)
        pdf_bytes = generate_pdf_from_html(html_content)
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=ap_aging_{as_of_date}.pdf"}
        )
    
    elif format == "excel":
        excel_bytes = generate_ap_aging_excel(report_data)
        return Response(
            content=excel_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=ap_aging_{as_of_date}.xlsx"}
        )

def generate_ap_aging_html(data: dict, org_settings: dict = None) -> str:
    """Generate HTML for AP Aging report"""
    org = org_settings or {}
    company_name = org.get('company_name', 'Battwheels')
    
    aging_data = data.get('aging_data', {})
    bills_html = ""
    for bill in data.get('bills', [])[:50]:
        bills_html += f"""
            <tr>
                <td>{bill.get('bill_number', '-')}</td>
                <td>{bill.get('vendor_name', '-')}</td>
                <td>{bill.get('due_date', '-')}</td>
                <td>{bill.get('days_overdue', 0)}</td>
                <td class="amount">{format_currency(bill.get('balance', 0))}</td>
            </tr>
        """
    
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            @page {{ size: A4 landscape; margin: 1cm; }}
            body {{ font-family: 'Helvetica Neue', Arial, sans-serif; font-size: 9pt; color: #333; padding: 20px; }}
            .header {{ text-align: center; border-bottom: 3px solid #22EDA9; padding-bottom: 15px; margin-bottom: 20px; }}
            .company-name {{ font-size: 18pt; font-weight: bold; }}
            .report-title {{ font-size: 14pt; color: #22EDA9; margin-top: 10px; }}
            .aging-summary {{ display: flex; justify-content: space-between; margin-bottom: 20px; gap: 10px; }}
            .aging-box {{ flex: 1; text-align: center; padding: 15px; background: #f8f9fa; border-radius: 8px; }}
            .aging-label {{ font-size: 9pt; color: #666; }}
            .aging-amount {{ font-size: 14pt; font-weight: bold; margin-top: 5px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
            th {{ background: #f8f9fa; padding: 10px 6px; text-align: left; border-bottom: 2px solid #22EDA9; }}
            td {{ padding: 8px 6px; border-bottom: 1px solid #eee; }}
            .amount {{ text-align: right; }}
            .grand-total {{ font-weight: bold; background: #22EDA9; }}
            .footer {{ margin-top: 20px; text-align: center; font-size: 8pt; color: #999; }}
        </style>
    </head>
    <body>
        <div class="header">
            <div class="company-name">{company_name}</div>
            <div class="report-title">ACCOUNTS PAYABLE AGING</div>
            <div style="font-size: 10pt; color: #666;">As of {data['as_of_date']}</div>
        </div>
        
        <div class="aging-summary">
            <div class="aging-box"><div class="aging-label">Current</div><div class="aging-amount">{format_currency(aging_data.get('current', 0))}</div></div>
            <div class="aging-box"><div class="aging-label">1-30 Days</div><div class="aging-amount">{format_currency(aging_data.get('1_30', 0))}</div></div>
            <div class="aging-box"><div class="aging-label">31-60 Days</div><div class="aging-amount">{format_currency(aging_data.get('31_60', 0))}</div></div>
            <div class="aging-box"><div class="aging-label">61-90 Days</div><div class="aging-amount">{format_currency(aging_data.get('61_90', 0))}</div></div>
            <div class="aging-box"><div class="aging-label">Over 90 Days</div><div class="aging-amount">{format_currency(aging_data.get('over_90', 0))}</div></div>
        </div>
        
        <table>
            <thead><tr><th>Bill #</th><th>Vendor</th><th>Due Date</th><th>Days Overdue</th><th class="amount">Balance</th></tr></thead>
            <tbody>
                {bills_html}
                <tr class="grand-total"><td colspan="4">Total Outstanding</td><td class="amount">{format_currency(data['total_ap'])}</td></tr>
            </tbody>
        </table>
        <div class="footer">Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')} | {company_name}</div>
    </body>
    </html>
    """
    return html

def generate_ap_aging_excel(data: dict) -> bytes:
    """Generate Excel for AP Aging report"""
    wb = create_styled_workbook()
    ws = wb.active
    ws.title = "AP Aging"
    
    ws.append(['Accounts Payable Aging Report'])
    ws.merge_cells('A1:E1')
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([f"As of: {data['as_of_date']}"])
    ws.append([])
    
    aging_data = data.get('aging_data', {})
    ws.append(['Aging Summary'])
    ws['A4'].font = Font(bold=True)
    ws.append(['Current', '1-30 Days', '31-60 Days', '61-90 Days', 'Over 90 Days'])
    style_header_row(ws, 5, 5)
    ws.append([aging_data.get('current', 0), aging_data.get('1_30', 0), aging_data.get('31_60', 0), aging_data.get('61_90', 0), aging_data.get('over_90', 0)])
    for col in range(1, 6):
        ws.cell(row=6, column=col).number_format = '₹#,##0.00'
    
    ws.append([])
    ws.append(['Bill Details'])
    ws['A8'].font = Font(bold=True)
    ws.append(['Bill Number', 'Vendor', 'Due Date', 'Days Overdue', 'Balance'])
    style_header_row(ws, 9, 5)
    
    for bill in data.get('bills', []):
        ws.append([bill.get('bill_number', '-'), bill.get('vendor_name', '-'), bill.get('due_date', '-'), bill.get('days_overdue', 0), bill.get('balance', 0)])
    
    ws.append(['Total', '', '', '', data['total_ap']])
    style_total_row(ws, ws.max_row, 5)
    
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    
    for row in ws.iter_rows(min_row=10, max_col=5):
        if row[4].value and isinstance(row[4].value, (int, float)):
            row[4].number_format = '₹#,##0.00'
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ============== TECHNICIAN PERFORMANCE REPORT ==============

@router.get("/technician-performance")
async def get_technician_performance(
    period: str = Query("this_month", description="this_week | this_month | this_quarter | custom"),
    date_from: str = Query(None),
    date_to: str = Query(None),
):
    """
    GET /api/reports/technician-performance
    Returns per-technician performance metrics for the specified period,
    ranked by a composite score: resolution_rate(0.4) + sla_compliance(0.4) + speed(0.2).
    """
    from fastapi import Request as _Req
    now = datetime.now(timezone.utc)

    # Determine date range
    if period == "this_week":
        start_dt = now - timedelta(days=7)
    elif period == "this_month":
        start_dt = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    elif period == "this_quarter":
        quarter_month = ((now.month - 1) // 3) * 3 + 1
        start_dt = now.replace(month=quarter_month, day=1, hour=0, minute=0, second=0, microsecond=0)
    elif period == "custom" and date_from:
        try:
            start_dt = datetime.fromisoformat(date_from.replace("Z", "+00:00"))
            if start_dt.tzinfo is None:
                start_dt = start_dt.replace(tzinfo=timezone.utc)
        except Exception:
            start_dt = now - timedelta(days=30)
    else:
        start_dt = now - timedelta(days=30)

    if period == "custom" and date_to:
        try:
            end_dt = datetime.fromisoformat(date_to.replace("Z", "+00:00"))
            if end_dt.tzinfo is None:
                end_dt = end_dt.replace(tzinfo=timezone.utc)
        except Exception:
            end_dt = now
    else:
        end_dt = now

    start_iso = start_dt.isoformat()
    end_iso = end_dt.isoformat()

    db = get_db()

    # Build query
    query = {
        "assigned_technician_id": {"$exists": True, "$ne": None},
        "created_at": {"$gte": start_iso, "$lte": end_iso}
    }

    tickets = await db.tickets.find(query, {
        "_id": 0,
        "ticket_id": 1,
        "assigned_technician_id": 1,
        "assigned_technician_name": 1,
        "status": 1,
        "priority": 1,
        "created_at": 1,
        "resolved_at": 1,
        "first_response_at": 1,
        "sla_response_due_at": 1,
        "sla_resolution_due_at": 1,
        "sla_response_breached": 1,
        "sla_resolution_breached": 1,
    }).to_list(5000)

    # Group by technician
    tech_map = {}
    for t in tickets:
        tid = t.get("assigned_technician_id")
        if not tid:
            continue
        if tid not in tech_map:
            tech_map[tid] = {
                "technician_id": tid,
                "technician_name": t.get("assigned_technician_name") or "Unknown",
                "tickets": []
            }
        tech_map[tid]["tickets"].append(t)

    # Also fetch technician names from users collection
    if tech_map:
        tech_users = await db.users.find(
            {"user_id": {"$in": list(tech_map.keys())}},
            {"_id": 0, "user_id": 1, "name": 1}
        ).to_list(200)
        for u in tech_users:
            if u["user_id"] in tech_map:
                tech_map[u["user_id"]]["technician_name"] = u.get("name") or tech_map[u["user_id"]]["technician_name"]

    resolved_statuses = {"resolved", "closed"}

    results = []
    for tid, data in tech_map.items():
        tkts = data["tickets"]
        total = len(tkts)
        resolved = [t for t in tkts if t.get("status") in resolved_statuses]
        total_resolved = len(resolved)
        resolution_rate = round((total_resolved / total * 100), 1) if total > 0 else 0.0

        # Response time (minutes)
        response_times = []
        for t in tkts:
            if t.get("first_response_at") and t.get("created_at"):
                try:
                    c = datetime.fromisoformat(t["created_at"].replace("Z", "+00:00"))
                    r = datetime.fromisoformat(t["first_response_at"].replace("Z", "+00:00"))
                    if c.tzinfo is None:
                        c = c.replace(tzinfo=timezone.utc)
                    if r.tzinfo is None:
                        r = r.replace(tzinfo=timezone.utc)
                    response_times.append((r - c).total_seconds() / 60)
                except Exception:
                    pass

        # Resolution time (minutes)
        resolution_times = []
        for t in resolved:
            if t.get("resolved_at") and t.get("created_at"):
                try:
                    c = datetime.fromisoformat(t["created_at"].replace("Z", "+00:00"))
                    r = datetime.fromisoformat(t["resolved_at"].replace("Z", "+00:00"))
                    if c.tzinfo is None:
                        c = c.replace(tzinfo=timezone.utc)
                    if r.tzinfo is None:
                        r = r.replace(tzinfo=timezone.utc)
                    resolution_times.append((r - c).total_seconds() / 60)
                except Exception:
                    pass

        avg_response = round(sum(response_times) / len(response_times), 1) if response_times else None
        avg_resolution = round(sum(resolution_times) / len(resolution_times), 1) if resolution_times else None

        # SLA metrics
        sla_breach_response = sum(1 for t in tkts if t.get("sla_response_breached"))
        sla_breach_resolution = sum(1 for t in tkts if t.get("sla_resolution_breached"))
        total_with_sla = sum(1 for t in tkts if t.get("sla_resolution_due_at"))
        sla_breached_total = sum(1 for t in tkts if t.get("sla_response_breached") or t.get("sla_resolution_breached"))
        within_sla = total_with_sla - sla_breached_total
        sla_compliance = round((within_sla / total_with_sla * 100), 1) if total_with_sla > 0 else 100.0

        # Ranking score: resolution_rate(0.4) + sla_compliance(0.4) + speed(0.2)
        speed_score = 0.0
        if avg_resolution and avg_resolution > 0:
            # Normalize: 1/time (lower time = higher score), capped at reasonable max
            speed_score = min(1.0, 480 / avg_resolution)  # 480 min = 8h benchmark
        score = (resolution_rate / 100 * 0.4) + (sla_compliance / 100 * 0.4) + (speed_score * 0.2)

        name = data["technician_name"]
        initials = "".join(w[0].upper() for w in name.split() if w)[:2]

        results.append({
            "technician_id": tid,
            "technician_name": name,
            "avatar_initials": initials,
            "total_tickets_assigned": total,
            "total_tickets_resolved": total_resolved,
            "resolution_rate_pct": resolution_rate,
            "avg_response_time_minutes": avg_response,
            "avg_resolution_time_minutes": avg_resolution,
            "sla_breaches_response": sla_breach_response,
            "sla_breaches_resolution": sla_breach_resolution,
            "sla_compliance_rate_pct": sla_compliance,
            "tickets_within_sla": within_sla,
            "customer_satisfaction_score": None,
            "_score": round(score, 4),
        })

    # Sort by score descending, assign rank
    results.sort(key=lambda x: x["_score"], reverse=True)
    for i, r in enumerate(results):
        r["rank"] = i + 1
        del r["_score"]

    return {
        "code": 0,
        "technicians": results,
        "period": period,
        "date_from": start_iso,
        "date_to": end_iso,
        "total_technicians": len(results),
    }

