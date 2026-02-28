"""
GST Compliance Module for Battwheels OS
Implements Indian GST (Goods and Services Tax) features:
- GSTIN validation
- CGST/SGST/IGST calculation based on place of supply
- GSTR-1 (Outward Supplies) report
- GSTR-3B (Summary Return) report
- HSN/SAC code management
- E-Invoice IRN generation (placeholder)
"""
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import Response
from datetime import datetime, timezone, timedelta
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, validator, Field
from typing import Optional, List
from io import BytesIO
import os
import re
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from fastapi import Request
from utils.database import extract_org_id, org_query


# Soft import for PDF service (may not be available in all environments)
PDF_SERVICE_AVAILABLE = False
generate_pdf_from_html = None
try:
    from services.pdf_service import generate_pdf_from_html
    PDF_SERVICE_AVAILABLE = True
except Exception:
    pass

router = APIRouter(prefix="/gst", tags=["GST Compliance"])

# Database connection
def get_db():
    mongo_url = os.environ['MONGO_URL']
    client = AsyncIOMotorClient(mongo_url)
    return client[os.environ['DB_NAME']]

# ============== INDIAN STATES ==============
INDIAN_STATES = {
    "01": "Jammu and Kashmir",
    "02": "Himachal Pradesh",
    "03": "Punjab",
    "04": "Chandigarh",
    "05": "Uttarakhand",
    "06": "Haryana",
    "07": "Delhi",
    "08": "Rajasthan",
    "09": "Uttar Pradesh",
    "10": "Bihar",
    "11": "Sikkim",
    "12": "Arunachal Pradesh",
    "13": "Nagaland",
    "14": "Manipur",
    "15": "Mizoram",
    "16": "Tripura",
    "17": "Meghalaya",
    "18": "Assam",
    "19": "West Bengal",
    "20": "Jharkhand",
    "21": "Odisha",
    "22": "Chhattisgarh",
    "23": "Madhya Pradesh",
    "24": "Gujarat",
    "25": "Daman and Diu",
    "26": "Dadra and Nagar Haveli",
    "27": "Maharashtra",
    "28": "Andhra Pradesh",
    "29": "Karnataka",
    "30": "Goa",
    "31": "Lakshadweep",
    "32": "Kerala",
    "33": "Tamil Nadu",
    "34": "Puducherry",
    "35": "Andaman and Nicobar Islands",
    "36": "Telangana",
    "37": "Andhra Pradesh (New)",
    "38": "Ladakh",
    "97": "Other Territory"
}

# State code mapping (short codes to GST codes)
STATE_CODE_MAP = {
    "JK": "01", "HP": "02", "PB": "03", "CH": "04", "UK": "05",
    "HR": "06", "DL": "07", "RJ": "08", "UP": "09", "BR": "10",
    "SK": "11", "AR": "12", "NL": "13", "MN": "14", "MZ": "15",
    "TR": "16", "ML": "17", "AS": "18", "WB": "19", "JH": "20",
    "OD": "21", "CG": "22", "MP": "23", "GJ": "24", "DD": "25",
    "DN": "26", "MH": "27", "AP": "28", "KA": "29", "GA": "30",
    "LD": "31", "KL": "32", "TN": "33", "PY": "34", "AN": "35",
    "TS": "36", "LA": "38"
}

# GST Rates
GST_RATES = [0, 5, 12, 18, 28]

# ============== GSTIN VALIDATION ==============
GSTIN_REGEX = r'^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[1-9A-Z]{1}Z[0-9A-Z]{1}$'


def compute_gstin_checksum(gstin_14: str) -> str:
    """
    Compute the checksum character (position 15) of a GSTIN.
    Uses mod-36 weighted sum algorithm as per GST specification.
    gstin_14: first 14 characters of GSTIN
    Returns: single checksum character (0-9 or A-Z)
    """
    CHARS = "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    factor = 1
    total = 0
    for char in gstin_14.upper():
        if char not in CHARS:
            return None  # Invalid character
        digit = CHARS.index(char)
        product = factor * digit
        product = (product // 36) + (product % 36)
        total += product
        factor = 2 if factor == 1 else 1
    check_digit = (36 - (total % 36)) % 36
    return CHARS[check_digit]


def validate_gstin(gstin: str) -> dict:
    """Validate GSTIN format, state code, and checksum (P2-18)"""
    if not gstin:
        return {"valid": False, "error": "GSTIN is empty"}
    
    gstin = gstin.upper().strip()
    
    if len(gstin) != 15:
        return {"valid": False, "error": "GSTIN must be 15 characters"}
    
    if not re.match(GSTIN_REGEX, gstin):
        return {"valid": False, "error": "Invalid GSTIN format"}
    
    state_code = gstin[:2]
    if state_code not in INDIAN_STATES:
        return {"valid": False, "error": f"Invalid state code: {state_code}"}
    
    # Checksum validation (P2-18) — mod-36 algorithm on first 14 chars
    expected_check = compute_gstin_checksum(gstin[:14])
    if expected_check is None or gstin[14] != expected_check:
        return {
            "valid": False,
            "error": f"Invalid GSTIN checksum. Expected {expected_check} at position 15."
        }
    
    # Extract details
    return {
        "valid": True,
        "gstin": gstin,
        "state_code": state_code,
        "state_name": INDIAN_STATES[state_code],
        "pan": gstin[2:12],
        "entity_code": gstin[12],
        "checksum": gstin[14]
    }

# ============== GST CALCULATION ==============
def calculate_gst(subtotal: float, gst_rate: float, org_state_code: str, customer_state_code: str) -> dict:
    """
    Calculate GST based on place of supply
    - Intra-state (same state): Split into CGST + SGST (half each)
    - Inter-state (different state): Full IGST
    """
    gst_amount = subtotal * (gst_rate / 100)
    
    is_intra_state = org_state_code == customer_state_code
    
    if is_intra_state:
        half = round(gst_amount / 2, 2)
        return {
            "is_intra_state": True,
            "gst_rate": gst_rate,
            "cgst_rate": gst_rate / 2,
            "cgst_amount": half,
            "sgst_rate": gst_rate / 2,
            "sgst_amount": half,
            "igst_rate": 0,
            "igst_amount": 0,
            "total_gst": round(gst_amount, 2)
        }
    else:
        return {
            "is_intra_state": False,
            "gst_rate": gst_rate,
            "cgst_rate": 0,
            "cgst_amount": 0,
            "sgst_rate": 0,
            "sgst_amount": 0,
            "igst_rate": gst_rate,
            "igst_amount": round(gst_amount, 2),
            "total_gst": round(gst_amount, 2)
        }

# ============== PYDANTIC MODELS ==============

class GSTINValidation(BaseModel):
    gstin: str

class OrganizationGSTSettings(BaseModel):
    gstin: str
    place_of_supply: str  # State code like "27" for Maharashtra
    legal_name: Optional[str] = ""
    trade_name: Optional[str] = ""
    address: Optional[str] = ""

class GSTCalculationRequest(BaseModel):
    subtotal: float
    gst_rate: float = 18
    org_state_code: str = "27"  # Default Maharashtra
    customer_state_code: str = "27"

# ============== API ENDPOINTS ==============

@router.get("/summary")
async def get_gst_summary(request: Request):
    org_id = extract_org_id(request)
    """Get GST summary for current financial year"""
    db = get_db()
    
    # Get current FY dates (April to March)
    today = datetime.now(timezone.utc)
    if today.month >= 4:
        fy_start = f"{today.year}-04-01"
        fy_end = f"{today.year + 1}-03-31"
    else:
        fy_start = f"{today.year - 1}-04-01"
        fy_end = f"{today.year}-03-31"
    
    # Get sales tax from invoices
    sales_pipeline = [
        {"$match": {"invoice_date": {"$gte": fy_start, "$lte": fy_end}, "status": {"$ne": "void"}}},
        {"$group": {
            "_id": None,
            "total_sales": {"$sum": "$sub_total"},
            "cgst": {"$sum": {"$ifNull": ["$cgst_amount", 0]}},
            "sgst": {"$sum": {"$ifNull": ["$sgst_amount", 0]}},
            "igst": {"$sum": {"$ifNull": ["$igst_amount", 0]}},
            "total_tax": {"$sum": "$tax_total"}
        }}
    ]
    sales = await db.invoices_enhanced.aggregate(sales_pipeline).to_list(1)
    
    # Get purchase tax from bills  
    purchase_pipeline = [
        {"$match": {"bill_date": {"$gte": fy_start, "$lte": fy_end}, "status": {"$ne": "void"}}},
        {"$group": {
            "_id": None,
            "total_purchases": {"$sum": "$sub_total"},
            "input_cgst": {"$sum": {"$ifNull": ["$cgst_amount", 0]}},
            "input_sgst": {"$sum": {"$ifNull": ["$sgst_amount", 0]}},
            "input_igst": {"$sum": {"$ifNull": ["$igst_amount", 0]}},
            "total_input_tax": {"$sum": "$tax_total"}
        }}
    ]
    purchases = await db.bills_enhanced.aggregate(purchase_pipeline).to_list(1)
    
    sales_data = sales[0] if sales else {}
    purchase_data = purchases[0] if purchases else {}
    
    output_tax = sales_data.get("total_tax", 0)
    input_tax = purchase_data.get("total_input_tax", 0)
    
    return {
        "code": 0,
        "summary": {
            "financial_year": f"{fy_start[:4]}-{fy_end[:4]}",
            "sales": {
                "taxable_value": round(sales_data.get("total_sales", 0), 2),
                "cgst": round(sales_data.get("cgst", 0), 2),
                "sgst": round(sales_data.get("sgst", 0), 2),
                "igst": round(sales_data.get("igst", 0), 2),
                "total_output_tax": round(output_tax, 2)
            },
            "purchases": {
                "taxable_value": round(purchase_data.get("total_purchases", 0), 2),
                "input_cgst": round(purchase_data.get("input_cgst", 0), 2),
                "input_sgst": round(purchase_data.get("input_sgst", 0), 2),
                "input_igst": round(purchase_data.get("input_igst", 0), 2),
                "total_input_tax": round(input_tax, 2)
            },
            "net_liability": round(output_tax - input_tax, 2)
        }
    }

@router.get("/states")
async def get_indian_states(request: Request):
    org_id = extract_org_id(request)
    """Get list of Indian states with GST codes"""
    states = [{"code": k, "name": v} for k, v in INDIAN_STATES.items()]
    return {"code": 0, "states": states}

@router.post("/validate-gstin")
async def validate_gstin_endpoint(request: Request, data: GSTINValidation):
    org_id = extract_org_id(request)
    """Validate GSTIN format and extract details"""
    result = validate_gstin(data.gstin)
    return {"code": 0 if result["valid"] else 1, **result}

@router.post("/calculate")
async def calculate_gst_endpoint(request: Request, data: GSTCalculationRequest):
    org_id = extract_org_id(request)
    """Calculate GST (CGST/SGST/IGST) based on place of supply"""
    result = calculate_gst(
        data.subtotal,
        data.gst_rate,
        data.org_state_code,
        data.customer_state_code
    )
    return {"code": 0, **result}

@router.get("/organization-settings")
async def get_organization_gst_settings(request: Request):
    org_id = extract_org_id(request)
    """Get organization GST settings"""
    db = get_db()
    settings = await db.organization_settings.find_one(
        org_query(org_id, {}), {"_id": 0}
    )
    if not settings:
        settings = {
            "gstin": "",
            "place_of_supply": "27",  # Default Maharashtra
            "legal_name": "Battwheels",
            "trade_name": "Battwheels"
        }
    return {"code": 0, "settings": settings}

@router.put("/organization-settings")
async def update_organization_gst_settings(request: Request, settings: OrganizationGSTSettings):
    org_id = extract_org_id(request)
    """Update organization GST settings"""
    db = get_db()
    
    # Validate GSTIN if provided
    if settings.gstin:
        validation = validate_gstin(settings.gstin)
        if not validation["valid"]:
            raise HTTPException(status_code=400, detail=validation["error"])
    
    await db.organization_settings.update_one(
        org_query(org_id, {}),
        {"$set": {
            "gstin": settings.gstin,
            "place_of_supply": settings.place_of_supply,
            "state_code": settings.place_of_supply,
            "legal_name": settings.legal_name,
            "trade_name": settings.trade_name,
            "address": settings.address
        }},
        upsert=True
    )
    return {"code": 0, "message": "GST settings updated"}

# ============== GSTR-1 REPORT ==============

@router.get("/gstr1")
async def get_gstr1_report(request: Request, month: str = "", # Format: YYYY-MM
    format: str = Query("json", enum=["json", "excel", "pdf"])
):
    """
    GSTR-1 Report: Outward Supplies (Sales Invoices)
    Categories:
    - B2B: Business to Business (with GSTIN)
    - B2C Large: B2C > 2.5L inter-state
    - B2C Small: B2C ≤ 2.5L
    - CDNR: Credit/Debit Notes (Registered)
    - EXP: Exports
    """
    db = get_db()
    
    # Default to current month
    if not month:
        month = datetime.now(timezone.utc).strftime("%Y-%m")
    
    try:
        year, mon = month.split("-")
        start_date = f"{year}-{mon}-01"
        # Get last day of month
        if int(mon) == 12:
            end_date = f"{int(year)+1}-01-01"
        else:
            end_date = f"{year}-{int(mon)+1:02d}-01"
    except:
        raise HTTPException(status_code=400, detail="Invalid month format. Use YYYY-MM")
    
    # Fetch invoices for the period — from invoices_enhanced (primary collection), org-scoped
    org_id = extract_org_id(request)
    inv_query = org_query(org_id, {
        "invoice_date": {"$gte": start_date, "$lt": end_date},
        "status": {"$in": ["sent", "paid", "partial", "overdue"]}
    })
    invoices = await db.invoices_enhanced.find(inv_query, {"_id": 0}).to_list(5000)  # SPRINT-2B: hard-cap unbounded query
    
    # Also fetch from legacy invoices collection (Zoho-synced)
    legacy_query = org_query(org_id, {
        "date": {"$gte": start_date, "$lt": end_date},
        "status": {"$in": ["sent", "paid", "partial", "overdue"]}
    })
    legacy_invoices = await db.invoices.find(legacy_query, {"_id": 0}).to_list(5000)  # SPRINT-2B: hard-cap unbounded query
    
    # Get organization settings — org-scoped
    org_settings = await db.organization_settings.find_one(
        org_query(org_id, {}), {"_id": 0}
    ) or {}
    org_state = org_settings.get("place_of_supply", "06")
    
    # Merge enhanced + legacy invoices
    all_invoices = list(invoices) + list(legacy_invoices)
    for inv in all_invoices:
        if "invoice_date" not in inv:
            inv["invoice_date"] = inv.get("date", "")
    
    # Categorize invoices
    b2b_invoices = []
    b2c_large = []
    b2c_small = []
    
    for inv in all_invoices:
        customer_gstin = inv.get("customer_gstin", "") or inv.get("gst_no", "")
        customer_state = inv.get("place_of_supply", org_state)
        
        # Extract state from GSTIN if available
        if customer_gstin and len(customer_gstin) >= 2:
            customer_state = customer_gstin[:2]
        
        is_intra_state = customer_state == org_state
        subtotal = inv.get("sub_total", 0) or inv.get("subtotal", 0) or 0
        total = inv.get("total", 0)
        tax_total = inv.get("tax_total", 0) or (total - subtotal)
        
        # Calculate CGST/SGST/IGST
        if is_intra_state:
            cgst = round(tax_total / 2, 2)
            sgst = round(tax_total / 2, 2)
            igst = 0
        else:
            cgst = 0
            sgst = 0
            igst = round(tax_total, 2)
        
        inv_data = {
            "invoice_number": inv.get("invoice_number", ""),
            "invoice_date": inv.get("date", ""),
            "customer_name": inv.get("customer_name", ""),
            "customer_gstin": customer_gstin,
            "place_of_supply": customer_state,
            "state_name": INDIAN_STATES.get(customer_state, "Unknown"),
            "taxable_value": subtotal,
            "cgst": cgst,
            "sgst": sgst,
            "igst": igst,
            "total_tax": tax_total,
            "invoice_value": total,
            "is_intra_state": is_intra_state
        }
        
        if customer_gstin and validate_gstin(customer_gstin)["valid"]:
            b2b_invoices.append(inv_data)
        elif total > 250000:
            b2c_large.append(inv_data)  # B2CL: no GSTIN, value > 2.5L
        else:
            b2c_small.append(inv_data)  # B2CS: no GSTIN, value <= 2.5L
    
    # Aggregate totals
    def aggregate(inv_list):
        return {
            "count": len(inv_list),
            "taxable_value": sum(i["taxable_value"] for i in inv_list),
            "cgst": sum(i["cgst"] for i in inv_list),
            "sgst": sum(i["sgst"] for i in inv_list),
            "igst": sum(i["igst"] for i in inv_list),
            "total_tax": sum(i["total_tax"] for i in inv_list),
            "invoice_value": sum(i["invoice_value"] for i in inv_list)
        }
    
    # Credit/Debit Notes for CDNR (Section 9A/34 of GSTR-1) — org-scoped
    cn_query = org_query(org_id, {
        "created_at": {"$gte": start_date, "$lt": end_date},
        "status": {"$ne": "cancelled"}
    })
    credit_notes = await db.credit_notes.find(cn_query, {"_id": 0}).to_list(1000)
    
    # Build CDNR entries with proper GST breakdown — split by registered/unregistered (P1-09)
    cdnr_entries = []
    cdnr_unregistered_entries = []
    for cn in credit_notes:
        cn_data = {
            "credit_note_number": cn.get("credit_note_number", ""),
            "credit_note_date": cn.get("created_at", "")[:10],
            "original_invoice_number": cn.get("original_invoice_number", ""),
            "customer_name": cn.get("customer_name", ""),
            "customer_gstin": cn.get("customer_gstin", ""),
            "reason": cn.get("reason", ""),
            "taxable_value": cn.get("subtotal", 0),
            "cgst": cn.get("cgst_amount", 0),
            "sgst": cn.get("sgst_amount", 0),
            "igst": cn.get("igst_amount", 0),
            "total_tax": cn.get("gst_amount", 0),
            "note_value": cn.get("total", 0),
            "note_type": "Credit Note",
            "gst_treatment": cn.get("gst_treatment", "cgst_sgst"),
        }
        buyer_gstin = cn.get("customer_gstin", "")
        if buyer_gstin and validate_gstin(buyer_gstin)["valid"]:
            cdnr_entries.append(cn_data)
        else:
            cdnr_unregistered_entries.append(cn_data)
    
    cdnr_summary = {
        "count": len(cdnr_entries),
        "taxable_value": sum(e["taxable_value"] for e in cdnr_entries),
        "cgst": sum(e["cgst"] for e in cdnr_entries),
        "sgst": sum(e["sgst"] for e in cdnr_entries),
        "igst": sum(e["igst"] for e in cdnr_entries),
        "total_tax": sum(e["total_tax"] for e in cdnr_entries),
        "total_value": sum(e["note_value"] for e in cdnr_entries),
    }
    
    cdnr_unreg_summary = {
        "count": len(cdnr_unregistered_entries),
        "taxable_value": sum(e["taxable_value"] for e in cdnr_unregistered_entries),
        "cgst": sum(e["cgst"] for e in cdnr_unregistered_entries),
        "sgst": sum(e["sgst"] for e in cdnr_unregistered_entries),
        "igst": sum(e["igst"] for e in cdnr_unregistered_entries),
        "total_tax": sum(e["total_tax"] for e in cdnr_unregistered_entries),
        "total_value": sum(e["note_value"] for e in cdnr_unregistered_entries),
    }
    
    # B2CS summary by state+rate (P1-09 — Table 7 GSTR-1)
    b2cs_by_state_rate = {}
    for inv in b2c_small:
        state = inv.get("state_name", "Unknown")
        rate = inv.get("gst_rate", 0)
        key = f"{state}_{rate}"
        if key not in b2cs_by_state_rate:
            b2cs_by_state_rate[key] = {
                "state": state,
                "gst_rate": rate,
                "taxable_value": 0, "cgst": 0, "sgst": 0, "igst": 0,
                "total_tax": 0, "count": 0
            }
        entry = b2cs_by_state_rate[key]
        entry["taxable_value"] += inv.get("taxable_value", 0)
        entry["cgst"] += inv.get("cgst", 0)
        entry["sgst"] += inv.get("sgst", 0)
        entry["igst"] += inv.get("igst", 0)
        entry["total_tax"] += inv.get("total_tax", 0)
        entry["count"] += 1
    
    # HSN summary by hsn_code + gst_rate (P1-09 — Table 12 GSTR-1)
    hsn_by_code_rate = {}
    all_invoices_for_hsn = b2b_invoices + b2c_large + b2c_small
    for inv in all_invoices_for_hsn:
        hsn = inv.get("hsn_code", inv.get("hsn_sac", "9987"))
        rate = inv.get("gst_rate", 0)
        key = f"{hsn}_{rate}"
        if key not in hsn_by_code_rate:
            hsn_by_code_rate[key] = {
                "hsn_code": hsn, "gst_rate": rate,
                "taxable_value": 0, "cgst": 0, "sgst": 0, "igst": 0,
                "total_quantity": 0, "uom": "NOS"
            }
        entry = hsn_by_code_rate[key]
        entry["taxable_value"] += inv.get("taxable_value", 0)
        entry["cgst"] += inv.get("cgst", 0)
        entry["sgst"] += inv.get("sgst", 0)
        entry["igst"] += inv.get("igst", 0)
        entry["total_quantity"] += 1
    
    # Net adjustments: grand_total should subtract credit note amounts
    cn_taxable = cdnr_summary["taxable_value"] + cdnr_unreg_summary["taxable_value"]
    cn_cgst = cdnr_summary["cgst"] + cdnr_unreg_summary["cgst"]
    cn_sgst = cdnr_summary["sgst"] + cdnr_unreg_summary["sgst"]
    cn_igst = cdnr_summary["igst"] + cdnr_unreg_summary["igst"]
    
    report_data = {
        "period": month,
        "filing_status": "draft",
        "b2b": {
            "summary": aggregate(b2b_invoices),
            "invoices": b2b_invoices
        },
        "b2c_large": {
            "summary": aggregate(b2c_large),
            "invoices": b2c_large
        },
        "b2c_small": {
            "summary": aggregate(b2c_small),
            "invoices": b2c_small,
            "by_state_rate": list(b2cs_by_state_rate.values())
        },
        "cdnr": {
            "summary": cdnr_summary,
            "notes": cdnr_entries
        },
        "cdnr_unregistered": {
            "summary": cdnr_unreg_summary,
            "notes": cdnr_unregistered_entries
        },
        "hsn_summary": {
            "by_code_rate": list(hsn_by_code_rate.values()),
            "totals": {
                "taxable_value": sum(h["taxable_value"] for h in hsn_by_code_rate.values()),
                "cgst": sum(h["cgst"] for h in hsn_by_code_rate.values()),
                "sgst": sum(h["sgst"] for h in hsn_by_code_rate.values()),
                "igst": sum(h["igst"] for h in hsn_by_code_rate.values()),
            }
        },
        "grand_total": {
            "taxable_value": aggregate(b2b_invoices)["taxable_value"] + aggregate(b2c_large)["taxable_value"] + aggregate(b2c_small)["taxable_value"] - cn_taxable,
            "cgst": aggregate(b2b_invoices)["cgst"] + aggregate(b2c_large)["cgst"] + aggregate(b2c_small)["cgst"] - cn_cgst,
            "sgst": aggregate(b2b_invoices)["sgst"] + aggregate(b2c_large)["sgst"] + aggregate(b2c_small)["sgst"] - cn_sgst,
            "igst": aggregate(b2b_invoices)["igst"] + aggregate(b2c_large)["igst"] + aggregate(b2c_small)["igst"] - cn_igst,
            "total_invoices": len(b2b_invoices) + len(b2c_large) + len(b2c_small),
            "total_credit_notes": len(cdnr_entries) + len(cdnr_unregistered_entries)
        }
    }
    
    if format == "excel":
        return generate_gstr1_excel(report_data, month)
    elif format == "pdf":
        return generate_gstr1_pdf(report_data, month, org_settings)
    
    return {"code": 0, "report": "gstr1", **report_data}

def generate_gstr1_excel(data: dict, month: str) -> Response:
    """Generate GSTR-1 Excel export"""
    wb = openpyxl.Workbook()
    
    # B2B Sheet
    ws_b2b = wb.active
    ws_b2b.title = "B2B"
    ws_b2b.append(["GSTR-1 B2B Invoices", f"Period: {month}"])
    ws_b2b.append([])
    ws_b2b.append(["Invoice No", "Date", "Customer", "GSTIN", "Place of Supply", "Taxable Value", "CGST", "SGST", "IGST", "Total"])
    for inv in data["b2b"]["invoices"]:
        ws_b2b.append([
            inv["invoice_number"], inv["invoice_date"], inv["customer_name"],
            inv["customer_gstin"], inv["state_name"],
            inv["taxable_value"], inv["cgst"], inv["sgst"], inv["igst"], inv["invoice_value"]
        ])
    ws_b2b.append(["Total", "", "", "", "",
        data["b2b"]["summary"]["taxable_value"],
        data["b2b"]["summary"]["cgst"],
        data["b2b"]["summary"]["sgst"],
        data["b2b"]["summary"]["igst"],
        data["b2b"]["summary"]["invoice_value"]
    ])
    
    # B2C Large Sheet
    ws_b2c = wb.create_sheet("B2C Large")
    ws_b2c.append(["GSTR-1 B2C Large Invoices (>2.5L Inter-state)"])
    ws_b2c.append([])
    ws_b2c.append(["Invoice No", "Date", "Customer", "Place of Supply", "Taxable Value", "IGST", "Total"])
    for inv in data["b2c_large"]["invoices"]:
        ws_b2c.append([inv["invoice_number"], inv["invoice_date"], inv["customer_name"],
                       inv["state_name"], inv["taxable_value"], inv["igst"], inv["invoice_value"]])
    
    # B2C Small Sheet
    ws_b2c_small = wb.create_sheet("B2C Small")
    ws_b2c_small.append(["GSTR-1 B2C Small Invoices"])
    ws_b2c_small.append([])
    ws_b2c_small.append(["Invoice No", "Date", "Customer", "Taxable Value", "CGST", "SGST", "Total"])
    for inv in data["b2c_small"]["invoices"][:100]:  # Limit for Excel
        ws_b2c_small.append([inv["invoice_number"], inv["invoice_date"], inv["customer_name"],
                            inv["taxable_value"], inv["cgst"], inv["sgst"], inv["invoice_value"]])
    
    # Summary Sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary.append(["GSTR-1 Summary", f"Period: {month}"])
    ws_summary.append([])
    ws_summary.append(["Category", "Count", "Taxable Value", "CGST", "SGST", "IGST", "Total Tax"])
    ws_summary.append(["B2B", data["b2b"]["summary"]["count"], data["b2b"]["summary"]["taxable_value"],
                       data["b2b"]["summary"]["cgst"], data["b2b"]["summary"]["sgst"],
                       data["b2b"]["summary"]["igst"], data["b2b"]["summary"]["total_tax"]])
    ws_summary.append(["B2C Large", data["b2c_large"]["summary"]["count"], data["b2c_large"]["summary"]["taxable_value"],
                       data["b2c_large"]["summary"]["cgst"], data["b2c_large"]["summary"]["sgst"],
                       data["b2c_large"]["summary"]["igst"], data["b2c_large"]["summary"]["total_tax"]])
    ws_summary.append(["B2C Small", data["b2c_small"]["summary"]["count"], data["b2c_small"]["summary"]["taxable_value"],
                       data["b2c_small"]["summary"]["cgst"], data["b2c_small"]["summary"]["sgst"],
                       data["b2c_small"]["summary"]["igst"], data["b2c_small"]["summary"]["total_tax"]])
    ws_summary.append(["Grand Total", data["grand_total"]["total_invoices"], data["grand_total"]["taxable_value"],
                       data["grand_total"]["cgst"], data["grand_total"]["sgst"],
                       data["grand_total"]["igst"], data["grand_total"]["cgst"] + data["grand_total"]["sgst"] + data["grand_total"]["igst"]])
    
    # Apply formatting
    for ws in wb.worksheets:
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 20
        ws['A1'].font = Font(bold=True, size=14)
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=gstr1_{month}.xlsx"}
    )

def generate_gstr1_pdf(data: dict, month: str, org_settings: dict) -> Response:
    """Generate GSTR-1 PDF report"""
    company = org_settings.get("company_name", "Battwheels")
    gstin = org_settings.get("gstin", "")
    
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            @page {{ size: A4 landscape; margin: 1cm; }}
            body {{ font-family: Arial, sans-serif; font-size: 9pt; }}
            .header {{ text-align: center; border-bottom: 2px solid #22EDA9; padding-bottom: 10px; margin-bottom: 15px; }}
            .company {{ font-size: 16pt; font-weight: bold; }}
            .title {{ font-size: 14pt; color: #22EDA9; margin-top: 5px; }}
            .gstin {{ font-size: 10pt; color: #666; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 15px; font-size: 8pt; }}
            th {{ background: #f0f0f0; padding: 8px 4px; text-align: left; border: 1px solid #ddd; }}
            td {{ padding: 6px 4px; border: 1px solid #ddd; }}
            .amount {{ text-align: right; }}
            .total-row {{ background: #22EDA9; font-weight: bold; }}
            .section {{ margin-top: 20px; }}
            .section-title {{ font-size: 12pt; font-weight: bold; color: #333; border-bottom: 1px solid #22EDA9; padding-bottom: 5px; }}
            .summary-box {{ display: inline-block; width: 23%; text-align: center; padding: 10px; background: #f8f8f8; margin: 5px; border-radius: 5px; }}
            .summary-label {{ font-size: 8pt; color: #666; }}
            .summary-value {{ font-size: 14pt; font-weight: bold; }}
        </style>
    </head>
    <body>
        <div class="header">
            <div class="company">{company}</div>
            <div class="title">GSTR-1 - Outward Supplies</div>
            <div class="gstin">GSTIN: {gstin} | Period: {month}</div>
        </div>
        
        <div class="section">
            <div class="section-title">Summary</div>
            <div style="margin-top: 10px;">
                <div class="summary-box"><div class="summary-label">Total Invoices</div><div class="summary-value">{data['grand_total']['total_invoices']}</div></div>
                <div class="summary-box"><div class="summary-label">Taxable Value</div><div class="summary-value">₹{data['grand_total']['taxable_value']:,.0f}</div></div>
                <div class="summary-box"><div class="summary-label">CGST+SGST</div><div class="summary-value">₹{(data['grand_total']['cgst'] + data['grand_total']['sgst']):,.0f}</div></div>
                <div class="summary-box"><div class="summary-label">IGST</div><div class="summary-value">₹{data['grand_total']['igst']:,.0f}</div></div>
            </div>
        </div>
        
        <div class="section">
            <div class="section-title">B2B Invoices ({data['b2b']['summary']['count']})</div>
            <table>
                <tr><th>Invoice No</th><th>Date</th><th>Customer</th><th>GSTIN</th><th class="amount">Taxable</th><th class="amount">CGST</th><th class="amount">SGST</th><th class="amount">IGST</th><th class="amount">Total</th></tr>
                {"".join(f"<tr><td>{i['invoice_number']}</td><td>{i['invoice_date']}</td><td>{i['customer_name'][:25]}</td><td>{i['customer_gstin']}</td><td class='amount'>₹{i['taxable_value']:,.0f}</td><td class='amount'>₹{i['cgst']:,.0f}</td><td class='amount'>₹{i['sgst']:,.0f}</td><td class='amount'>₹{i['igst']:,.0f}</td><td class='amount'>₹{i['invoice_value']:,.0f}</td></tr>" for i in data['b2b']['invoices'][:30])}
                <tr class="total-row"><td colspan="4">Total B2B</td><td class="amount">₹{data['b2b']['summary']['taxable_value']:,.0f}</td><td class="amount">₹{data['b2b']['summary']['cgst']:,.0f}</td><td class="amount">₹{data['b2b']['summary']['sgst']:,.0f}</td><td class="amount">₹{data['b2b']['summary']['igst']:,.0f}</td><td class="amount">₹{data['b2b']['summary']['invoice_value']:,.0f}</td></tr>
            </table>
        </div>
        
        <div class="section">
            <div class="section-title">B2C Summary</div>
            <table>
                <tr><th>Category</th><th class="amount">Count</th><th class="amount">Taxable Value</th><th class="amount">CGST</th><th class="amount">SGST</th><th class="amount">IGST</th></tr>
                <tr><td>B2C Large (>2.5L Inter-state)</td><td class="amount">{data['b2c_large']['summary']['count']}</td><td class="amount">₹{data['b2c_large']['summary']['taxable_value']:,.0f}</td><td class="amount">₹{data['b2c_large']['summary']['cgst']:,.0f}</td><td class="amount">₹{data['b2c_large']['summary']['sgst']:,.0f}</td><td class="amount">₹{data['b2c_large']['summary']['igst']:,.0f}</td></tr>
                <tr><td>B2C Small</td><td class="amount">{data['b2c_small']['summary']['count']}</td><td class="amount">₹{data['b2c_small']['summary']['taxable_value']:,.0f}</td><td class="amount">₹{data['b2c_small']['summary']['cgst']:,.0f}</td><td class="amount">₹{data['b2c_small']['summary']['sgst']:,.0f}</td><td class="amount">₹{data['b2c_small']['summary']['igst']:,.0f}</td></tr>
            </table>
        </div>
        
        <div style="margin-top: 20px; text-align: center; font-size: 8pt; color: #999;">
            Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')} | This is a computer-generated document
        </div>
    </body>
    </html>
    """
    
    pdf_bytes = generate_pdf_from_html(html)
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=gstr1_{month}.pdf"}
    )

# ============== GSTR-3B REPORT ==============

@router.get("/gstr3b")
async def get_gstr3b_report(request: Request, month: str = "", # Format: YYYY-MM
    format: str = Query("json", enum=["json", "excel", "pdf"])
):
    """
    GSTR-3B Report: Summary Return
    - 3.1: Outward supplies (from invoices)
    - 3.2: Interstate supplies
    - 4: Input Tax Credit (from bills/expenses)
    - 5.1: Interest and late fees
    - 6.1: Payment of tax
    """
    db = get_db()
    
    if not month:
        month = datetime.now(timezone.utc).strftime("%Y-%m")
    
    try:
        year, mon = month.split("-")
        start_date = f"{year}-{mon}-01"
        if int(mon) == 12:
            end_date = f"{int(year)+1}-01-01"
        else:
            end_date = f"{year}-{int(mon)+1:02d}-01"
    except:
        raise HTTPException(status_code=400, detail="Invalid month format. Use YYYY-MM")
    
    # Extract org context — every query in this function MUST include org_id
    org_id = extract_org_id(request)
    
    # Get organization settings — org-scoped
    org_settings = await db.organization_settings.find_one(
        org_query(org_id, {}), {"_id": 0}
    ) or {}
    org_state = org_settings.get("place_of_supply", "27")
    
    # OUTWARD SUPPLIES (Invoices) — org-scoped from both collections
    inv_query = org_query(org_id, {
        "invoice_date": {"$gte": start_date, "$lt": end_date},
        "status": {"$in": ["sent", "paid", "partial", "overdue"]}
    })
    invoices_enh = await db.invoices_enhanced.find(inv_query, {"_id": 0}).to_list(5000)  # SPRINT-2B: hard-cap unbounded query
    
    legacy_query = org_query(org_id, {
        "date": {"$gte": start_date, "$lt": end_date},
        "status": {"$in": ["sent", "paid", "partial", "overdue"]}
    })
    invoices_leg = await db.invoices.find(legacy_query, {"_id": 0}).to_list(5000)  # SPRINT-2B: hard-cap unbounded query
    all_3b_invoices = list(invoices_enh) + list(invoices_leg)
    
    outward_taxable = 0
    outward_cgst = 0
    outward_sgst = 0
    outward_igst = 0
    
    # B2C tracking for Section 3.2 (P1-10: inter-state vs intra-state split)
    b2c_interstate_taxable = 0
    b2c_interstate_igst = 0
    b2c_interstate_by_state = {}
    b2c_intrastate_taxable = 0
    b2c_intrastate_cgst = 0
    b2c_intrastate_sgst = 0
    
    for inv in all_3b_invoices:
        subtotal = inv.get("sub_total", 0) or inv.get("subtotal", 0) or 0
        tax_total = inv.get("tax_total", 0) or (inv.get("total", 0) - subtotal)
        customer_gstin = inv.get("customer_gstin", "") or inv.get("gst_no", "")
        customer_state = customer_gstin[:2] if customer_gstin and len(customer_gstin) >= 2 else org_state
        is_intra = customer_state == org_state
        
        outward_taxable += subtotal
        
        if is_intra:
            outward_cgst += tax_total / 2
            outward_sgst += tax_total / 2
        else:
            outward_igst += tax_total
        
        # B2C split (Section 3.2 / Table 3.2) — invoices without valid GSTIN
        is_b2c = not (customer_gstin and validate_gstin(customer_gstin).get("valid"))
        if is_b2c:
            if not is_intra:
                b2c_interstate_taxable += subtotal
                b2c_interstate_igst += tax_total
                state_name = INDIAN_STATES.get(customer_state, "Unknown")
                if customer_state not in b2c_interstate_by_state:
                    b2c_interstate_by_state[customer_state] = {
                        "place_of_supply": customer_state,
                        "state_name": state_name,
                        "taxable_value": 0, "igst": 0
                    }
                b2c_interstate_by_state[customer_state]["taxable_value"] += subtotal
                b2c_interstate_by_state[customer_state]["igst"] += tax_total
            else:
                b2c_intrastate_taxable += subtotal
                b2c_intrastate_cgst += tax_total / 2
                b2c_intrastate_sgst += tax_total / 2
    
    # INPUT TAX CREDIT (Bills + Expenses) — org-scoped
    bills = await db.bills.find(
        org_query(org_id, {"date": {"$gte": start_date, "$lt": end_date}}),
        {"_id": 0}
    ).to_list(5000)  # SPRINT-2B: hard-cap unbounded query
    
    expenses = await db.expenses.find(
        org_query(org_id, {"expense_date": {"$gte": start_date, "$lt": end_date}}),
        {"_id": 0}
    ).to_list(5000)  # SPRINT-2B: hard-cap unbounded query
    
    input_taxable = 0
    input_cgst = 0
    input_sgst = 0
    input_igst = 0
    
    # ITC categorization for Section 4 / Table 4 (P1-10)
    itc_import_goods = {"cgst": 0, "sgst": 0, "igst": 0}
    itc_import_services = {"cgst": 0, "sgst": 0, "igst": 0}
    itc_rcm = {"cgst": 0, "sgst": 0, "igst": 0}
    itc_isd = {"cgst": 0, "sgst": 0, "igst": 0}
    itc_all_other = {"cgst": 0, "sgst": 0, "igst": 0}
    
    for bill in bills:
        subtotal = bill.get("sub_total", 0) or bill.get("subtotal", 0) or 0
        tax_total = bill.get("tax_total", 0) or (bill.get("total", 0) - subtotal)
        vendor_gstin = bill.get("vendor_gstin", "") or bill.get("gst_no", "")
        vendor_state = vendor_gstin[:2] if vendor_gstin and len(vendor_gstin) >= 2 else org_state
        
        input_taxable += subtotal
        
        if vendor_state == org_state:
            bill_cgst = tax_total / 2
            bill_sgst = tax_total / 2
            bill_igst = 0
        else:
            bill_cgst = 0
            bill_sgst = 0
            bill_igst = tax_total
        
        input_cgst += bill_cgst
        input_sgst += bill_sgst
        input_igst += bill_igst
        
        # Categorize ITC (Table 4A — P1-10)
        gst_treatment = bill.get("gst_treatment", "")
        is_import = bill.get("is_import", False) or gst_treatment in ("import", "overseas")
        is_rcm = bill.get("reverse_charge", False)
        is_isd = bill.get("is_isd", False) or gst_treatment == "isd"
        is_service = bill.get("is_service", False) or bill.get("supply_type") == "service"
        
        if is_import and is_service:
            target = itc_import_services
        elif is_import:
            target = itc_import_goods
        elif is_rcm:
            target = itc_rcm
        elif is_isd:
            target = itc_isd
        else:
            target = itc_all_other
        
        target["cgst"] += bill_cgst
        target["sgst"] += bill_sgst
        target["igst"] += bill_igst
    
    for exp in expenses:
        amount = exp.get("amount", 0)
        tax_amount = exp.get("tax_amount", 0) or amount * 0.18  # Assume 18% if not specified
        exp_cgst = tax_amount / 2
        exp_sgst = tax_amount / 2
        input_taxable += amount
        input_cgst += exp_cgst
        input_sgst += exp_sgst
        itc_all_other["cgst"] += exp_cgst
        itc_all_other["sgst"] += exp_sgst
    
    # TABLE 4B — ITC Reversed (Sprint 4B-04)
    # vendor_credits schema: {amount, line_items: [{rate, tax_rate, amount}], status, date, organization_id}
    # No cgst_amount/sgst_amount/igst_amount fields — compute from line_items and linked bill state
    vendor_credit_query = org_query(org_id, {
        "date": {"$gte": start_date, "$lte": end_date},
        "status": {"$in": ["approved", "issued", "applied"]}
    })
    vendor_credits_list = await db.vendor_credits.find(
        vendor_credit_query, {"_id": 0}
    ).to_list(1000)

    itc_reversed_cgst = 0
    itc_reversed_sgst = 0
    itc_reversed_igst = 0
    for vc in vendor_credits_list:
        # Sum tax from line_items
        vc_tax = 0
        for li in vc.get("line_items", []):
            item_amount = li.get("amount", 0) or (li.get("quantity", 0) * li.get("rate", 0))
            item_tax_rate = li.get("tax_rate", 0) or 0
            vc_tax += item_amount * (item_tax_rate / 100)
        # If no line_items, estimate from total amount with zero tax
        # Determine intra/inter from linked bill's vendor GSTIN
        linked_bill_id = vc.get("linked_bill_id")
        vendor_state = org_state  # default to intra-state
        if linked_bill_id:
            linked_bill = await db.bills.find_one(
                {"bill_id": linked_bill_id, "organization_id": org_id},
                {"vendor_gstin": 1, "_id": 0}
            )
            if linked_bill:
                vg = linked_bill.get("vendor_gstin", "")
                if vg and len(vg) >= 2:
                    vendor_state = vg[:2]
        if vendor_state == org_state:
            itc_reversed_cgst += vc_tax / 2
            itc_reversed_sgst += vc_tax / 2
        else:
            itc_reversed_igst += vc_tax

    itc_reversed_others = {
        "cgst": itc_reversed_cgst,
        "sgst": itc_reversed_sgst,
        "igst": itc_reversed_igst,
    }

    # Rule 42/43: ITC reversal for exempt/non-business use
    # ITC must be reversed proportional to exempt supply ratio.
    # Formula: Reversal = Total_ITC_Available × (Exempt_Supply / Total_Supply)
    # If an org has NO exempt supplies (most EV workshops), the ratio is 0
    # and the reversal is correctly zero — this is not a gap, it's the
    # expected result under Rule 42/43 when 100% of supplies are taxable.
    exempt_invoices_query = org_query(org_id, {
        "invoice_date": {"$gte": start_date, "$lt": end_date},
        "supply_type": "exempt"
    })
    exempt_invoices = await db.invoices_enhanced.find(
        exempt_invoices_query, {"_id": 0, "sub_total": 1}
    ).to_list(5000)
    total_exempt_value = sum(
        inv.get("sub_total", 0) for inv in exempt_invoices
    )
    total_supply_value = outward_taxable  # total taxable outward (before CN)

    if total_exempt_value > 0 and total_supply_value > 0:
        exempt_ratio = total_exempt_value / (total_supply_value + total_exempt_value)
        total_itc_available_cgst = input_cgst
        total_itc_available_sgst = input_sgst
        total_itc_available_igst = input_igst
        itc_reversed_rule42_43 = {
            "cgst": round(total_itc_available_cgst * exempt_ratio, 2),
            "sgst": round(total_itc_available_sgst * exempt_ratio, 2),
            "igst": round(total_itc_available_igst * exempt_ratio, 2),
        }
    else:
        # No exempt supplies — Rule 42/43 reversal is correctly zero.
        # This is the expected result for businesses with 100% taxable supplies.
        itc_reversed_rule42_43 = {"cgst": 0, "sgst": 0, "igst": 0}

    itc_reversed_total_cgst = itc_reversed_rule42_43["cgst"] + itc_reversed_others["cgst"]
    itc_reversed_total_sgst = itc_reversed_rule42_43["sgst"] + itc_reversed_others["sgst"]
    itc_reversed_total_igst = itc_reversed_rule42_43["igst"] + itc_reversed_others["igst"]

    # TABLE 4D — Ineligible ITC (Sprint 4B-04)
    # Query bills where is_blocked_credit = True (Section 17(5) blocked inputs)
    blocked_bills_query = org_query(org_id, {
        "date": {"$gte": start_date, "$lt": end_date},
        "is_blocked_credit": True
    })
    blocked_bills = await db.bills.find(
        blocked_bills_query, {"_id": 0}
    ).to_list(1000)
    # NOTE: Requires is_blocked_credit flag on bill documents.
    # Currently zero if flag not set — correct default.
    # Bill schema: {sub_total, tax_total, total, vendor_gstin} — compute CGST/SGST/IGST from state

    itc_ineligible_cgst = 0
    itc_ineligible_sgst = 0
    itc_ineligible_igst = 0
    for b in blocked_bills:
        b_tax = b.get("tax_total", 0) or (b.get("total", 0) - (b.get("sub_total", 0) or b.get("subtotal", 0) or 0))
        b_vendor_gstin = b.get("vendor_gstin", "") or b.get("gst_no", "")
        b_vendor_state = b_vendor_gstin[:2] if b_vendor_gstin and len(b_vendor_gstin) >= 2 else org_state
        if b_vendor_state == org_state:
            itc_ineligible_cgst += b_tax / 2
            itc_ineligible_sgst += b_tax / 2
        else:
            itc_ineligible_igst += b_tax

    itc_ineligible_17_5 = {
        "cgst": itc_ineligible_cgst,
        "sgst": itc_ineligible_sgst,
        "igst": itc_ineligible_igst,
    }
    itc_ineligible_others = {"cgst": 0, "sgst": 0, "igst": 0}

    itc_ineligible_total_cgst = itc_ineligible_17_5["cgst"] + itc_ineligible_others["cgst"]
    itc_ineligible_total_sgst = itc_ineligible_17_5["sgst"] + itc_ineligible_others["sgst"]
    itc_ineligible_total_igst = itc_ineligible_17_5["igst"] + itc_ineligible_others["igst"]

    # NET ITC = 4A total - 4B reversed - 4D ineligible
    net_itc_cgst = input_cgst - itc_reversed_total_cgst - itc_ineligible_total_cgst
    net_itc_sgst = input_sgst - itc_reversed_total_sgst - itc_ineligible_total_sgst
    net_itc_igst = input_igst - itc_reversed_total_igst - itc_ineligible_total_igst

    # NET TAX LIABILITY (uses net ITC after reversals and ineligible deductions)
    net_cgst = max(0, outward_cgst - net_itc_cgst)
    net_sgst = max(0, outward_sgst - net_itc_sgst)
    net_igst = max(0, outward_igst - net_itc_igst)
    
    # Credit Notes (reduce output liability) — org-scoped, proper GST breakdown
    cn_query_3b = org_query(org_id, {
        "created_at": {"$gte": start_date, "$lt": end_date},
        "status": {"$ne": "cancelled"}
    })
    cn_list = await db.credit_notes.find(cn_query_3b, {"_id": 0}).to_list(1000)
    cn_taxable = sum(cn.get("subtotal", 0) for cn in cn_list)
    cn_cgst = sum(cn.get("cgst_amount", 0) for cn in cn_list)
    cn_sgst = sum(cn.get("sgst_amount", 0) for cn in cn_list)
    cn_igst = sum(cn.get("igst_amount", 0) for cn in cn_list)
    cn_tax_total = cn_cgst + cn_sgst + cn_igst
    cn_value = sum(cn.get("total", 0) for cn in cn_list)
    
    # Adjust net liability by deducting CN tax
    net_cgst = max(0, outward_cgst - net_itc_cgst - cn_cgst)
    net_sgst = max(0, outward_sgst - net_itc_sgst - cn_sgst)
    net_igst = max(0, outward_igst - net_itc_igst - cn_igst)
    
    # SECTION 3.1(d) — Inward supplies liable to reverse charge — org-scoped
    rcm_bill_query = org_query(org_id, {
        "date": {"$gte": start_date, "$lt": end_date},
        "reverse_charge": True
    })
    rcm_bills = await db.bills.find(rcm_bill_query, {"_id": 0}).to_list(5000)  # SPRINT-2B: hard-cap unbounded query
    
    rcm_taxable = 0
    rcm_cgst = 0
    rcm_sgst = 0
    rcm_igst = 0
    
    for bill in rcm_bills:
        subtotal = bill.get("sub_total", 0) or bill.get("subtotal", 0) or 0
        tax_total = bill.get("tax_total", 0) or (bill.get("total", 0) - subtotal)
        vendor_gstin = bill.get("vendor_gstin", "") or bill.get("gst_no", "")
        vendor_state = vendor_gstin[:2] if vendor_gstin and len(vendor_gstin) >= 2 else org_state
        
        rcm_taxable += subtotal
        if vendor_state == org_state:
            rcm_cgst += tax_total / 2
            rcm_sgst += tax_total / 2
        else:
            rcm_igst += tax_total
    
    rcm_total_tax = rcm_cgst + rcm_sgst + rcm_igst
    
    # Add RCM to net tax liability (RCM is payable ON TOP of forward charge)
    net_cgst += rcm_cgst
    net_sgst += rcm_sgst
    net_igst += rcm_igst
    
    report_data = {
        "period": month,
        "filing_status": "draft",
        "section_3_1": {
            "description": "Outward taxable supplies (net of credit notes)",
            "taxable_value": round(outward_taxable - cn_taxable, 2),
            "cgst": round(outward_cgst - cn_cgst, 2),
            "sgst": round(outward_sgst - cn_sgst, 2),
            "igst": round(outward_igst - cn_igst, 2),
            "total_tax": round(outward_cgst + outward_sgst + outward_igst - cn_tax_total, 2),
            "gross_outward": round(outward_taxable, 2),
            "cn_adjustment": round(cn_taxable, 2)
        },
        "section_3_1_d": {
            "description": "Inward supplies liable to reverse charge",
            "taxable_value": round(rcm_taxable, 2),
            "cgst": round(rcm_cgst, 2),
            "sgst": round(rcm_sgst, 2),
            "igst": round(rcm_igst, 2),
            "total_tax": round(rcm_total_tax, 2),
            "bill_count": len(rcm_bills)
        },
        "section_3_2": {
            "description": "Inter-state supplies to unregistered persons (Table 3.2)",
            "interstate": {
                "supplies": list(b2c_interstate_by_state.values()),
                "total_taxable_value": round(b2c_interstate_taxable, 2),
                "total_igst": round(b2c_interstate_igst, 2),
            },
            "intrastate": {
                "total_taxable_value": round(b2c_intrastate_taxable, 2),
                "total_cgst": round(b2c_intrastate_cgst, 2),
                "total_sgst": round(b2c_intrastate_sgst, 2),
            }
        },
        "section_4": {
            "description": "Eligible ITC (Input Tax Credit — Table 4)",
            "table_4A": {
                "description": "ITC Available (whether in full or part)",
                "(1)_import_of_goods": {"cgst": round(itc_import_goods["cgst"], 2), "sgst": round(itc_import_goods["sgst"], 2), "igst": round(itc_import_goods["igst"], 2)},
                "(2)_import_of_services": {"cgst": round(itc_import_services["cgst"], 2), "sgst": round(itc_import_services["sgst"], 2), "igst": round(itc_import_services["igst"], 2)},
                "(3)_inward_supplies_rcm": {"cgst": round(itc_rcm["cgst"], 2), "sgst": round(itc_rcm["sgst"], 2), "igst": round(itc_rcm["igst"], 2)},
                "(4)_inward_supplies_isd": {"cgst": round(itc_isd["cgst"], 2), "sgst": round(itc_isd["sgst"], 2), "igst": round(itc_isd["igst"], 2)},
                "(5)_all_other_itc": {"cgst": round(itc_all_other["cgst"], 2), "sgst": round(itc_all_other["sgst"], 2), "igst": round(itc_all_other["igst"], 2)},
            },
            "table_4B": {
                "description": "ITC Reversed",
                "(1)_as_per_rules_42_43": {"cgst": round(itc_reversed_rule42_43["cgst"], 2), "sgst": round(itc_reversed_rule42_43["sgst"], 2), "igst": round(itc_reversed_rule42_43["igst"], 2)},
                "(2)_others": {"cgst": round(itc_reversed_others["cgst"], 2), "sgst": round(itc_reversed_others["sgst"], 2), "igst": round(itc_reversed_others["igst"], 2)},
            },
            "table_4C": {
                "description": "Net ITC Available (A) - (B)",
                "cgst": round(net_itc_cgst, 2),
                "sgst": round(net_itc_sgst, 2),
                "igst": round(net_itc_igst, 2),
                "total_itc": round(net_itc_cgst + net_itc_sgst + net_itc_igst, 2)
            },
            "table_4D": {
                "description": "Ineligible ITC",
                "(1)_as_per_section_17_5": {"cgst": round(itc_ineligible_17_5["cgst"], 2), "sgst": round(itc_ineligible_17_5["sgst"], 2), "igst": round(itc_ineligible_17_5["igst"], 2)},
                "(2)_others": {"cgst": round(itc_ineligible_others["cgst"], 2), "sgst": round(itc_ineligible_others["sgst"], 2), "igst": round(itc_ineligible_others["igst"], 2)},
            }
        },
        "section_5": {
            "description": "Exempt, Nil-rated & Non-GST",
            "inter_state": 0,
            "intra_state": 0
        },
        "section_6": {
            "description": "Payment of Tax",
            "net_cgst": round(net_cgst, 2),
            "net_sgst": round(net_sgst, 2),
            "net_igst": round(net_igst, 2),
            "total_liability": round(net_cgst + net_sgst + net_igst, 2),
            "interest": 0,
            "late_fee": 0
        },
        "adjustments": {
            "credit_notes": {
                "count": len(cn_list),
                "taxable_value": round(cn_taxable, 2),
                "cgst": round(cn_cgst, 2),
                "sgst": round(cn_sgst, 2),
                "igst": round(cn_igst, 2),
                "total_tax": round(cn_tax_total, 2),
                "total_value": round(cn_value, 2)
            }
        },
        "summary": {
            "total_output_tax": round(outward_cgst + outward_sgst + outward_igst - cn_tax_total, 2),
            "total_input_tax": round(net_itc_cgst + net_itc_sgst + net_itc_igst, 2),
            "itc_reversed": round(itc_reversed_total_cgst + itc_reversed_total_sgst + itc_reversed_total_igst, 2),
            "itc_ineligible": round(itc_ineligible_total_cgst + itc_ineligible_total_sgst + itc_ineligible_total_igst, 2),
            "rcm_tax_liability": round(rcm_total_tax, 2),
            "net_tax_payable": round(net_cgst + net_sgst + net_igst, 2)
        }
    }
    
    if format == "excel":
        return generate_gstr3b_excel(report_data, month)
    elif format == "pdf":
        return generate_gstr3b_pdf(report_data, month, org_settings)
    
    return {"code": 0, "report": "gstr3b", **report_data}

def generate_gstr3b_excel(data: dict, month: str) -> Response:
    """Generate GSTR-3B Excel export"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GSTR-3B"
    
    ws.append(["GSTR-3B Summary Return", f"Period: {month}"])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([])
    
    # Section 3.1 - Outward Supplies
    ws.append(["3.1 Outward Supplies"])
    ws['A3'].font = Font(bold=True)
    ws.append(["Description", "Taxable Value", "CGST", "SGST", "IGST"])
    ws.append(["Outward taxable supplies", data["section_3_1"]["taxable_value"],
               data["section_3_1"]["cgst"], data["section_3_1"]["sgst"], data["section_3_1"]["igst"]])
    ws.append(["(d) Inward supplies (reverse charge)", data["section_3_1_d"]["taxable_value"],
               data["section_3_1_d"]["cgst"], data["section_3_1_d"]["sgst"], data["section_3_1_d"]["igst"]])
    ws.append([])
    
    # Section 4 - ITC
    ws.append(["4. Eligible ITC"])
    ws['A7'].font = Font(bold=True)
    ws.append(["Description", "CGST", "SGST", "IGST"])
    ws.append(["Input Tax Credit", data["section_4"]["table_4C"]["cgst"], data["section_4"]["table_4C"]["sgst"], data["section_4"]["table_4C"]["igst"]])
    ws.append([])
    
    # Section 6 - Payment
    ws.append(["6.1 Payment of Tax"])
    ws['A11'].font = Font(bold=True)
    ws.append(["Tax", "Liability", "ITC Available", "Net Payable"])
    ws.append(["CGST", data["section_3_1"]["cgst"], data["section_4"]["table_4C"]["cgst"], data["section_6"]["net_cgst"]])
    ws.append(["SGST", data["section_3_1"]["sgst"], data["section_4"]["table_4C"]["sgst"], data["section_6"]["net_sgst"]])
    ws.append(["IGST", data["section_3_1"]["igst"], data["section_4"]["table_4C"]["igst"], data["section_6"]["net_igst"]])
    ws.append(["Total", data["section_3_1"]["total_tax"], data["section_4"]["table_4C"]["total_itc"], data["section_6"]["total_liability"]])
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=gstr3b_{month}.xlsx"}
    )

def generate_gstr3b_pdf(data: dict, month: str, org_settings: dict) -> Response:
    """Generate GSTR-3B PDF report"""
    company = org_settings.get("company_name", "Battwheels")
    gstin = org_settings.get("gstin", "")
    
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            @page {{ size: A4; margin: 1cm; }}
            body {{ font-family: Arial, sans-serif; font-size: 10pt; }}
            .header {{ text-align: center; border-bottom: 2px solid #22EDA9; padding-bottom: 10px; margin-bottom: 15px; }}
            .company {{ font-size: 16pt; font-weight: bold; }}
            .title {{ font-size: 14pt; color: #22EDA9; margin-top: 5px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
            th {{ background: #f0f0f0; padding: 8px; text-align: left; border: 1px solid #ddd; }}
            td {{ padding: 8px; border: 1px solid #ddd; }}
            .amount {{ text-align: right; }}
            .section {{ margin-top: 20px; }}
            .section-title {{ font-size: 12pt; font-weight: bold; background: #f8f8f8; padding: 8px; border-left: 4px solid #22EDA9; }}
            .total {{ font-weight: bold; background: #22EDA9; }}
            .summary-box {{ border: 2px solid #22EDA9; padding: 15px; margin-top: 20px; background: #f8fff8; }}
        </style>
    </head>
    <body>
        <div class="header">
            <div class="company">{company}</div>
            <div class="title">GSTR-3B - Summary Return</div>
            <div style="font-size: 10pt; color: #666;">GSTIN: {gstin} | Period: {month}</div>
        </div>
        
        <div class="section">
            <div class="section-title">3.1 Outward Supplies (Taxable)</div>
            <table>
                <tr><th>Description</th><th class="amount">Taxable Value</th><th class="amount">CGST</th><th class="amount">SGST</th><th class="amount">IGST</th></tr>
                <tr><td>Outward taxable supplies</td><td class="amount">₹{data['section_3_1']['taxable_value']:,.2f}</td><td class="amount">₹{data['section_3_1']['cgst']:,.2f}</td><td class="amount">₹{data['section_3_1']['sgst']:,.2f}</td><td class="amount">₹{data['section_3_1']['igst']:,.2f}</td></tr>
                <tr><td>(d) Inward supplies (reverse charge)</td><td class="amount">₹{data['section_3_1_d']['taxable_value']:,.2f}</td><td class="amount">₹{data['section_3_1_d']['cgst']:,.2f}</td><td class="amount">₹{data['section_3_1_d']['sgst']:,.2f}</td><td class="amount">₹{data['section_3_1_d']['igst']:,.2f}</td></tr>
            </table>
        </div>
        
        <div class="section">
            <div class="section-title">4. Eligible ITC (Input Tax Credit)</div>
            <table>
                <tr><th>Description</th><th class="amount">CGST</th><th class="amount">SGST</th><th class="amount">IGST</th><th class="amount">Total</th></tr>
                <tr><td>Input Tax Credit</td><td class="amount">₹{data['section_4']['table_4C']['cgst']:,.2f}</td><td class="amount">₹{data['section_4']['table_4C']['sgst']:,.2f}</td><td class="amount">₹{data['section_4']['table_4C']['igst']:,.2f}</td><td class="amount">₹{data['section_4']['table_4C']['total_itc']:,.2f}</td></tr>
            </table>
        </div>
        
        <div class="section">
            <div class="section-title">6.1 Payment of Tax</div>
            <table>
                <tr><th>Tax</th><th class="amount">Output Tax</th><th class="amount">ITC Used</th><th class="amount">Net Payable</th></tr>
                <tr><td>CGST</td><td class="amount">₹{data['section_3_1']['cgst']:,.2f}</td><td class="amount">₹{data['section_4']['table_4C']['cgst']:,.2f}</td><td class="amount">₹{data['section_6']['net_cgst']:,.2f}</td></tr>
                <tr><td>SGST</td><td class="amount">₹{data['section_3_1']['sgst']:,.2f}</td><td class="amount">₹{data['section_4']['table_4C']['sgst']:,.2f}</td><td class="amount">₹{data['section_6']['net_sgst']:,.2f}</td></tr>
                <tr><td>IGST</td><td class="amount">₹{data['section_3_1']['igst']:,.2f}</td><td class="amount">₹{data['section_4']['table_4C']['igst']:,.2f}</td><td class="amount">₹{data['section_6']['net_igst']:,.2f}</td></tr>
                <tr class="total"><td>Total</td><td class="amount">₹{data['section_3_1']['total_tax']:,.2f}</td><td class="amount">₹{data['section_4']['table_4C']['total_itc']:,.2f}</td><td class="amount">₹{data['section_6']['total_liability']:,.2f}</td></tr>
            </table>
        </div>
        
        <div class="summary-box">
            <table style="border: none;">
                <tr><td style="border: none;"><strong>Total Output Tax:</strong></td><td style="border: none; text-align: right;">₹{data['summary']['total_output_tax']:,.2f}</td></tr>
                <tr><td style="border: none;"><strong>Total Input Tax (ITC):</strong></td><td style="border: none; text-align: right;">₹{data['summary']['total_input_tax']:,.2f}</td></tr>
                <tr><td style="border: none; font-size: 14pt;"><strong>NET TAX PAYABLE:</strong></td><td style="border: none; text-align: right; font-size: 14pt; color: #22EDA9;"><strong>₹{data['summary']['net_tax_payable']:,.2f}</strong></td></tr>
            </table>
        </div>
        
        <div style="margin-top: 20px; text-align: center; font-size: 8pt; color: #999;">
            Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')} | This is a computer-generated document
        </div>
    </body>
    </html>
    """
    
    pdf_bytes = generate_pdf_from_html(html)
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=gstr3b_{month}.pdf"}
    )

# ============== HSN SUMMARY ==============

@router.get("/hsn-summary")
async def get_hsn_summary(request: Request, month: str = "", format: str = Query("json", enum=["json", "excel"])
):
    """
    HSN-wise Summary of Outward Supplies
    Required for GSTR-1 filing
    """
    db = get_db()
    org_id = extract_org_id(request)
    
    if not month:
        month = datetime.now(timezone.utc).strftime("%Y-%m")
    
    try:
        year, mon = month.split("-")
        start_date = f"{year}-{mon}-01"
        if int(mon) == 12:
            end_date = f"{int(year)+1}-01-01"
        else:
            end_date = f"{year}-{int(mon)+1:02d}-01"
    except:
        raise HTTPException(status_code=400, detail="Invalid month format")
    
    # Get all invoice line items with HSN codes — org-scoped (Sprint 2C carry-forward)
    invoices = await db.invoices.find(
        {"organization_id": org_id, "date": {"$gte": start_date, "$lt": end_date}},
        {"_id": 0, "line_items": 1}
    ).to_list(5000)  # SPRINT-2B: hard-cap unbounded query
    
    hsn_data = {}
    
    for inv in invoices:
        for item in inv.get("line_items", []):
            hsn = item.get("hsn_or_sac", "") or item.get("hsn_code", "") or "N/A"
            if hsn not in hsn_data:
                hsn_data[hsn] = {
                    "hsn_code": hsn,
                    "description": item.get("description", "") or item.get("name", ""),
                    "uqc": item.get("unit", "NOS"),
                    "quantity": 0,
                    "taxable_value": 0,
                    "cgst": 0,
                    "sgst": 0,
                    "igst": 0
                }
            
            qty = item.get("quantity", 1)
            rate = item.get("rate", 0)
            tax_rate = item.get("tax_percentage", 18) or 18
            amount = qty * rate
            tax = amount * (tax_rate / 100)
            
            hsn_data[hsn]["quantity"] += qty
            hsn_data[hsn]["taxable_value"] += amount
            hsn_data[hsn]["cgst"] += tax / 2
            hsn_data[hsn]["sgst"] += tax / 2
    
    hsn_list = list(hsn_data.values())
    
    if format == "excel":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "HSN Summary"
        ws.append(["HSN Summary", f"Period: {month}"])
        ws.append([])
        ws.append(["HSN Code", "Description", "UQC", "Quantity", "Taxable Value", "CGST", "SGST", "IGST"])
        for h in hsn_list:
            ws.append([h["hsn_code"], h["description"][:30], h["uqc"], h["quantity"],
                      round(h["taxable_value"], 2), round(h["cgst"], 2), round(h["sgst"], 2), round(h["igst"], 2)])
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return Response(
            content=buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=hsn_summary_{month}.xlsx"}
        )
    
    return {
        "code": 0,
        "period": month,
        "hsn_summary": hsn_list,
        "total": {
            "taxable_value": sum(h["taxable_value"] for h in hsn_list),
            "cgst": sum(h["cgst"] for h in hsn_list),
            "sgst": sum(h["sgst"] for h in hsn_list),
            "igst": sum(h["igst"] for h in hsn_list)
        }
    }
